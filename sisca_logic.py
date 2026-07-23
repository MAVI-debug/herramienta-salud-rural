"""
sisca_logic.py — Lógica de extracción PDF y generación SISCA
==============================================================
Independiente de PyQt6; reutilizable desde Flask y desde CLI.
"""

import os, re, math, copy, gc, unicodedata
from datetime import date
from io import BytesIO
from typing import Optional

import pypdf
import openpyxl
import openpyxl.utils
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.drawing.image import Image as XLImage, _import_image

# ---------------------------------------------------------------------------
# Constantes SISCA
# ---------------------------------------------------------------------------
SISCA_HOJA_ADELANTE = "NOMINAL adelante"
SISCA_HOJA_ATRAS    = "NOMINAL atras"
SISCA_FILAS_POR_HOJA = 15
SISCA_ALUMNOS_POR_BLOQUE = 30
SISCA_FILA_INICIO_ADELANTE = 19
SISCA_FILA_INICIO_ATRAS = 7

COLUMNAS_SUBTOTAL_SISCA = ("M", "N", "T")
FILA_SUBTOTAL_ATRAS = 6
FILA_TOTAL_ATRAS = 22

TIPO_CENTRO_DEFECTO = "PUBLICO"

# ---------------------------------------------------------------------------
#  Utilidades de texto
# ---------------------------------------------------------------------------

def _quitar_acentos(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

def _normalizar(texto: str) -> str:
    return _quitar_acentos(str(texto or "")).strip().lower()

def _nombre_archivo_seguro(texto: str) -> str:
    limpio = _quitar_acentos(str(texto or "SIN_NOMBRE"))
    limpio = re.sub(r'[\\/*?:"<>|]', "", limpio)
    limpio = re.sub(r"\s+", "_", limpio.strip())
    return limpio[:80] if limpio else "ESCUELA"

def _split_fecha(fecha: str):
    try:
        p = str(fecha).split("/")
        return int(p[0]), int(p[1]), int(p[2])
    except Exception:
        return None, None, None

def _nombre_completo(alumno: dict) -> str:
    return f"{alumno['nombres']} {alumno['apellidos']}"

def calcular_edad_a_fecha_corte(dia_nac, mes_nac, anio_nac, fecha_corte: date):
    try:
        dia_nac, mes_nac, anio_nac = int(dia_nac), int(mes_nac), int(anio_nac)
        edad = fecha_corte.year - anio_nac
        if (fecha_corte.month, fecha_corte.day) < (mes_nac, dia_nac):
            edad -= 1
        return edad
    except Exception:
        return None

# ---------------------------------------------------------------------------
#  Extracción del PDF
# ---------------------------------------------------------------------------

def extraer_alumnos_pdf(ruta_pdf: str) -> list:
    alumnos = []
    grado_actual = ""
    seccion_actual = ""
    _errores_pagina = []
    _cuis_vistos = set()

    _RE_CUI = re.compile(r'\b(\d{13})\b')
    _RE_CODIGO_PERSONAL = re.compile(r'\b([A-Z]{1,4}[\-]?[0-9]{3,8})\b')
    _RE_DATE = re.compile(r'(\d{1,2}/\d{1,2}/\d{4})')
    _RE_GEN = re.compile(r'\b(FEMENINO|FEM|MASCULINO|MASC|[FM])\b', re.IGNORECASE)
    _RE_PALABRA = re.compile(r'[A-Za-zÁÉÍÓÚÑáéíóúñ][A-Za-zÁÉÍÓÚÑáéíóúñ\.\-]+')
    _JUNK = {
        'GUATEMALTECA', 'GUATEMALA', 'CUI', 'GENERO', 'GENÉRO', 'SEXO',
        'FECHA', 'NACIMIENTO', 'NAC', 'LUGAR', 'NOMBRE', 'NOMBRES',
        'APELLIDOS', 'APELLIDO', 'GRADO', 'SECCION', 'SECCIÓN',
        'EDAD', 'TOTAL', 'TOTALES', 'PAIS', 'PAÍS', 'MUNICIPAL',
        'DEPTO', 'DEPARTAMENTO', 'DISTRITO', 'MODULO', 'MÓDULO',
        'CODIGO', 'CÓDIGO', 'ESCUELA', 'ESTABLECIMIENTO', 'INSTITUTO',
        'COMUNIDAD', 'CASERIO', 'ALDEA', 'COLONIA', 'BARRIO',
        'PLANILLA', 'NOMINAL', 'LISTADO', 'MINISTERIO', 'EDUCACION',
        'EDUCACIÓN', 'REPUBLICA', 'REPÚBLICA', 'SALUD', 'JORNADA',
        'MATUTINA', 'VESPERTINA', 'PRIMERA', 'SEGUNDA', 'DESCARGADO',
        'SISTEMA', 'NORTE', 'SUR', 'ESTE', 'OESTE', 'M', 'F',
    }

    def _es_linea_header(linea):
        ln = linea.strip().upper()
        if not ln or len(ln) < 3:
            return True
        skips = [
            'TOTAL', 'TOTALES', 'RESUMEN', 'PAGINA', 'PÁGINA',
            'FECHA', 'HORA', 'DIRECTOR', 'ENCARGADO', 'MINISTERIO',
            'REPUBLICA', 'REPÚBLICA', 'EDUCACION', 'EDUCACIÓN',
            'JORNADA', 'PLANILLA', 'NOMINAL', 'LISTADO',
            'APELLIDOS', 'NOMBRES', 'NOMBRE', 'APELLIDO',
            'FECHA DE NAC', 'LUGAR', 'CUI', 'GENERO', 'GENÉRO',
            'GRADO', 'SECCION', 'SECCIÓN', 'SEXO',
            'CODIGO', 'CÓDIGO', 'ESCUELA', 'DIRECCION', 'DIRECCIÓN',
            'MUNICIPIO', 'DEPARTAMENTO', 'DISTRITO',
            'No.', 'NO.', 'Nº', '#',
            'TOTAL GENERAL', 'TOTAL DE', 'SUBTOTAL',
            'EDAD', 'EDADES', '<=5', '6-9', '10-14', '15-19',
            'DESCARGADO', 'SISTEMA', 'MODULO', 'MÓDULO',
        ]
        for s in skips:
            if ln.startswith(s):
                return True
        if re.match(r'^\s*\d+\s*$', ln):
            return True
        if _RE_CUI.search(ln):
            return False
        if _RE_CODIGO_PERSONAL.search(ln):
            return False
        return False

    def _extraer_alumno_de_texto(bloque):
        m_cui = _RE_CUI.search(bloque)
        m_cod = _RE_CODIGO_PERSONAL.search(bloque) if not m_cui else None
        if not m_cui and not m_cod:
            return None

        if m_cui:
            cui = m_cui.group(1)
            if cui in _cuis_vistos:
                return None
            _cuis_vistos.add(cui)
        else:
            cui = ""

        m_gen = _RE_GEN.search(bloque)
        gen_raw = m_gen.group(1).upper() if m_gen else ""
        if gen_raw.startswith("F"):
            gen = "F"
        elif gen_raw.startswith("M"):
            gen = "M"
        else:
            gen = ""

        m_fec = _RE_DATE.search(bloque)
        fec = m_fec.group(1) if m_fec else ""

        texto_nombre = bloque[:m_cui.start() if m_cui else m_cod.start()]
        texto_nombre = re.sub(r'\b\d{1,3}\b', ' ', texto_nombre)
        texto_nombre = re.sub(r'\d{1,2}/\d{1,2}/\d{4}', ' ', texto_nombre)
        texto_nombre = re.sub(r'\b[A-Z]{1,4}[\-]?[0-9]{3,8}\b', ' ', texto_nombre)
        texto_nombre = re.sub(r'\s+', ' ', texto_nombre).strip()

        palabras = _RE_PALABRA.findall(texto_nombre)
        palabras = [p.upper() for p in palabras if p.upper() not in _JUNK and len(p) >= 2]

        ap = ""
        nom = ""
        if len(palabras) >= 4:
            mitad = len(palabras) // 2
            ap = " ".join(palabras[:mitad])
            nom = " ".join(palabras[mitad:])
        elif len(palabras) == 3:
            ap = " ".join(palabras[:2])
            nom = palabras[2]
        elif len(palabras) == 2:
            ap = palabras[0]
            nom = palabras[1]
        elif len(palabras) == 1:
            ap = palabras[0]

        if ap and nom:
            return {
                "grado": grado_actual,
                "seccion": seccion_actual,
                "apellidos": ap,
                "nombres": nom,
                "fecha_nac": fec,
                "cui": cui,
                "genero": gen,
            }
        return None

    def _procesar_pagina_texto(texto):
        nonlocal grado_actual, seccion_actual

        m_g = re.search(r'Grado:\s*(.+?)(?:\s+Seccion:|$)', texto, re.IGNORECASE)
        m_s = re.search(r'Seccion:\s*(\S+)', texto, re.IGNORECASE)
        if m_g:
            val = m_g.group(1).strip().upper()
            if val and len(val) < 40:
                grado_actual = val
        if m_s:
            val = m_s.group(1).strip().upper()
            if val:
                seccion_actual = val

        lineas = texto.splitlines()
        i = 0
        while i < len(lineas):
            linea = lineas[i]
            tiene_cui = bool(_RE_CUI.search(linea))
            tiene_codigo = bool(_RE_CODIGO_PERSONAL.search(linea))

            if not tiene_cui and not tiene_codigo:
                i += 1
                continue

            if _es_linea_header(linea):
                i += 1
                continue

            bloque = linea.rstrip()
            j = i + 1
            while j < len(lineas) and j <= i + 3:
                prox = lineas[j].strip()
                if prox and not _es_linea_header(prox):
                    bloque += " " + prox
                    j += 1
                else:
                    break

            alumno = _extraer_alumno_de_texto(bloque)
            if alumno:
                alumnos.append(alumno)
            i = j

    try:
        reader = pypdf.PdfReader(ruta_pdf)
        total = len(reader.pages)
    except Exception as e:
        print(f"[extraer_alumnos_pdf] No se pudo abrir PDF: {e}")
        return alumnos

    for idx in range(total):
        texto = ""
        try:
            pagina = reader.pages[idx]
            texto = pagina.extract_text() or ""
            del pagina
            _procesar_pagina_texto(texto)
        except Exception as e_pag:
            _errores_pagina.append(f"Página {idx + 1}: {type(e_pag).__name__}")
        finally:
            del texto
            gc.collect()

    del reader

    if _errores_pagina:
        print(f"[extraer_alumnos_pdf] {len(_errores_pagina)} error(es): "
              + "; ".join(_errores_pagina[:5])
              + ("..." if len(_errores_pagina) > 5 else ""))

    return alumnos

_ETIQUETAS_ENCABEZADO_PDF = ["Nombre:", "Dirección:", "Código:"]
_MAX_LINEAS_ENCABEZADO = 15

def extraer_metadatos_encabezado_pdf(ruta_pdf: str):
    texto_pag1 = ""
    try:
        reader = pypdf.PdfReader(ruta_pdf)
        texto_pag1 = reader.pages[0].extract_text() or ""
        del reader
    except Exception:
        return "", "", ""
    finally:
        gc.collect()

    lineas = texto_pag1.splitlines()[:_MAX_LINEAS_ENCABEZADO]
    texto_top = "\n".join(lineas)

    def valor_tras_etiqueta(etiqueta, texto):
        idx = texto.find(etiqueta)
        if idx == -1:
            return ""
        resto = texto[idx + len(etiqueta):]
        cortes = [resto.find(e) for e in _ETIQUETAS_ENCABEZADO_PDF
                  if e != etiqueta and resto.find(e) != -1]
        idx_salto = resto.find("\n")
        if idx_salto != -1:
            cortes.append(idx_salto)
        fin = min(cortes) if cortes else len(resto)
        valor = resto[:fin].replace("\n", " ")
        return re.sub(r"\s+", " ", valor).strip()

    nombre = valor_tras_etiqueta("Nombre:", texto_top).upper()
    direccion = valor_tras_etiqueta("Dirección:", texto_top).upper()
    codigo_raw = valor_tras_etiqueta("Código:", texto_top).upper()
    m = re.search(r"\d{1,2}-\d{1,2}-\d{2,5}-\d{1,3}", codigo_raw)
    codigo = m.group(0) if m else codigo_raw
    jornada = valor_tras_etiqueta("Jornada:", texto_top).upper()
    if "MATUTINA" in jornada:
        nombre = f"{nombre} JM"
    elif "VESPERTINA" in jornada:
        nombre = f"{nombre} JV"
    del texto_pag1, texto_top
    gc.collect()
    return nombre, direccion, codigo

def construir_nombre_escolar_completo(nombre: str, direccion: str) -> str:
    combinado = f"{nombre.strip()} {direccion.strip()}".strip()
    combinado = combinado.replace("\n", " ")
    combinado = re.sub(r"\s+", " ", combinado)
    return combinado.upper()

# ---------------------------------------------------------------------------
#  Generación de la ficha SISCA (openpyxl)
# ---------------------------------------------------------------------------

def _aplicar_fuente_encabezado(celda):
    alineacion_previa = celda.alignment
    celda.font = Font(name="Arial", size=14, bold=True)
    celda.alignment = Alignment(
        horizontal=alineacion_previa.horizontal,
        vertical=alineacion_previa.vertical,
        wrap_text=alineacion_previa.wrap_text,
        text_rotation=alineacion_previa.text_rotation,
    )

def _clonar_imagenes(ws_origen, ws_destino, prefijo=""):
    imagenes_origen = getattr(ws_origen, "_images", None)
    if not imagenes_origen:
        return
    for img_original in imagenes_origen:
        try:
            pil_img = _import_image(img_original.ref)
            buf = BytesIO()
            fmt = getattr(pil_img, "format", None) or "PNG"
            pil_img.save(buf, format=fmt)
            buf.seek(0)
            nueva = XLImage(buf)
            nueva.width = img_original.width
            nueva.height = img_original.height
            nueva.anchor = copy.deepcopy(img_original.anchor)
            nombre_archivo = img_original.path.split("/")[-1]
            nueva._path = f"/xl/media/{prefijo}_{nombre_archivo}" if prefijo \
                          else f"/xl/media/clon_{id(nueva)}_{nombre_archivo}"
        except Exception:
            continue
        ws_destino.add_image(nueva)

def _fijar_formulas_subtotal_y_total(ws_atras, nombre_adelante: str):
    for col in COLUMNAS_SUBTOTAL_SISCA:
        celda_subtotal = f"{col}{FILA_SUBTOTAL_ATRAS}"
        formula_actual = ws_atras[celda_subtotal].value
        if isinstance(formula_actual, str) and formula_actual.startswith("="):
            if SISCA_HOJA_ADELANTE in formula_actual:
                ws_atras[celda_subtotal] = formula_actual.replace(
                    f"'{SISCA_HOJA_ADELANTE}'", f"'{nombre_adelante}'")
        else:
            ws_atras[celda_subtotal] = f"='{nombre_adelante}'!{celda_subtotal}"
        celda_total = f"{col}{FILA_TOTAL_ATRAS}"
        rango_x = f"{col}{SISCA_FILA_INICIO_ATRAS}:{col}{FILA_TOTAL_ATRAS - 1}"
        ws_atras[celda_total] = f'=COUNTIF({rango_x},"X")+{celda_subtotal}'

def _asegurar_sheet_properties(ws):
    if ws is not None and ws.sheet_properties is None:
        from openpyxl.worksheet.properties import WorksheetProperties
        ws.sheet_properties = WorksheetProperties()

def _duplicar_bloque_sisca(wb, indice_bloque: int):
    if indice_bloque == 1:
        return wb[SISCA_HOJA_ADELANTE], wb[SISCA_HOJA_ATRAS]

    nombre_adelante = f"{SISCA_HOJA_ADELANTE} {indice_bloque}"
    nombre_atras    = f"{SISCA_HOJA_ATRAS} {indice_bloque}"

    ws_adelante_base = wb[SISCA_HOJA_ADELANTE]
    ws_atras_base    = wb[SISCA_HOJA_ATRAS]

    _asegurar_sheet_properties(ws_adelante_base)
    _asegurar_sheet_properties(ws_atras_base)

    ws_adelante = wb.copy_worksheet(ws_adelante_base)
    ws_adelante.title = nombre_adelante
    _clonar_imagenes(ws_adelante_base, ws_adelante, f"bloq{indice_bloque}")

    ws_atras = wb.copy_worksheet(ws_atras_base)
    ws_atras.title = nombre_atras
    _clonar_imagenes(ws_atras_base, ws_atras, f"bloq{indice_bloque}")

    try:
        if ws_adelante_base and ws_adelante:
            ws_adelante.print_area = ws_adelante_base.print_area
            if hasattr(ws_adelante_base, 'page_setup') and ws_adelante_base.page_setup:
                ws_adelante.page_setup.orientation = ws_adelante_base.page_setup.orientation
                ws_adelante.page_setup.paperSize = ws_adelante_base.page_setup.paperSize
                ws_adelante.page_setup.fitToPage = ws_adelante_base.page_setup.fitToPage
    except Exception as e:
        print(f"Nota: No se pudieron copiar propiedades de impresión adelante: {e}")

    try:
        if ws_atras_base and ws_atras:
            ws_atras.print_area = ws_atras_base.print_area
            if hasattr(ws_atras_base, 'page_setup') and ws_atras_base.page_setup:
                ws_atras.page_setup.orientation = ws_atras_base.page_setup.orientation
                ws_atras.page_setup.paperSize = ws_atras_base.page_setup.paperSize
                ws_atras.page_setup.fitToPage = ws_atras_base.page_setup.fitToPage
    except Exception as e:
        print(f"Nota: No se pudieron copiar propiedades de impresión atrás: {e}")

    _fijar_formulas_subtotal_y_total(ws_atras, nombre_adelante)
    return ws_adelante, ws_atras

def _rellenar_encabezado_sisca(ws_adelante, nombre_escuela: str, codigo_escuela: str,
                                tipo_centro: str = None,
                                area=None, distrito=None,
                                servicio=None, responsable=None, cargo=None,
                                fecha_reporte_str=None,
                                jornada="Primera Jornada"):
    tipo = (tipo_centro or TIPO_CENTRO_DEFECTO).strip().upper()
    valores = {
        "C7":  area or "",
        "T7":  distrito or "",
        "AB7": servicio or "",
        "G9":  responsable or "",
        "S9":  cargo or "",
        "AI9": fecha_reporte_str or "",
        "AE9": "X" if jornada == "Primera Jornada" else "",
        "AG9": "X" if jornada == "Segunda Jornada" else "",
        "E12": nombre_escuela.strip().upper(),
        "V12": codigo_escuela.strip().upper(),
    }
    if tipo == "PRIVADO":
        valores["AK12"] = "X"
    else:
        valores["AH12"] = "X"

    for direccion, valor in valores.items():
        ws_adelante[direccion] = valor
        _aplicar_fuente_encabezado(ws_adelante[direccion])

def _rellenar_alumnos_pagina(ws, fila_inicio: int, alumnos_pagina: list):
    for i in range(SISCA_FILAS_POR_HOJA):
        fila = fila_inicio + i
        for col in (2, 12, 13, 14, 15, 16, 17):
            ws.cell(fila, col, None)
    for i, alumno in enumerate(alumnos_pagina):
        fila = fila_inicio + i
        ws.cell(fila, 2, alumno["nombre"])
        celda_cui = ws.cell(fila, 12, "" if str(alumno["cui"]).startswith("TMP-") else alumno["cui"])
        celda_cui.number_format = "@"
        if alumno["genero"] == "F":
            ws.cell(fila, 13, "X")
        elif alumno["genero"] == "M":
            ws.cell(fila, 14, "X")
        ws.cell(fila, 15, alumno["dia"])
        ws.cell(fila, 16, alumno["mes"])
        ws.cell(fila, 17, alumno["anio"])

def generar_ficha_sisca_escuela(ruta_plantilla: str, ruta_salida: str,
                                 nombre_escuela: str, codigo_escuela: str,
                                 alumnos_aptos: list, anio_campana: int,
                                 tipo_centro: str = None,
                                 responsable=None, cargo=None,
                                 area=None, distrito=None, servicio=None,
                                 fecha_reporte_str=None,
                                 jornada="Primera Jornada") -> str:
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(
            f"No se encontró la plantilla legal en:\n{ruta_plantilla}")
    wb = openpyxl.load_workbook(ruta_plantilla)

    total_bloques = max(1, math.ceil(len(alumnos_aptos) / SISCA_ALUMNOS_POR_BLOQUE))
    bloques_hojas = [_duplicar_bloque_sisca(wb, i) for i in range(1, total_bloques + 1)]

    for idx, (ws_adelante, ws_atras) in enumerate(bloques_hojas, start=1):
        _rellenar_encabezado_sisca(ws_adelante, nombre_escuela, codigo_escuela,
                                    tipo_centro=tipo_centro,
                                    area=area, distrito=distrito,
                                    servicio=servicio,
                                    responsable=responsable, cargo=cargo,
                                    fecha_reporte_str=fecha_reporte_str,
                                    jornada=jornada)
        inicio = (idx - 1) * SISCA_ALUMNOS_POR_BLOQUE
        bloque = alumnos_aptos[inicio:inicio + SISCA_ALUMNOS_POR_BLOQUE]
        _rellenar_alumnos_pagina(ws_adelante, SISCA_FILA_INICIO_ADELANTE,
                                 bloque[:SISCA_FILAS_POR_HOJA])
        _rellenar_alumnos_pagina(ws_atras, SISCA_FILA_INICIO_ATRAS,
                                 bloque[SISCA_FILAS_POR_HOJA:SISCA_ALUMNOS_POR_BLOQUE])

    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass
    wb.save(ruta_salida)
    wb.close()
    gc.collect()
    return ruta_salida


def procesar_pdf_sisca(ruta_pdf: str, ruta_plantilla: str, ruta_salida_dir: str,
                       responsable: str, cargo: str, area_salud: str,
                       distrito_salud: str, servicio_salud: str,
                       tipo_centro: str, fecha_reporte_str: str = None,
                       fecha_corte: Optional[date] = None) -> dict:
    """
    Flujo completo:
      1. Extraer alumnos y metadatos del PDF
      2. Filtrar por edad (6-14 años) según fecha_corte
      3. Generar la ficha SISCA en Excel
      4. Retornar dict con ruta, metadatos de escuela y lista completa de alumnos
    """
    if not fecha_reporte_str:
        hoy = date.today()
        fecha_reporte_str = f"{hoy.day:02d}/{hoy.month:02d}/{hoy.year}"

    if fecha_corte is None:
        fecha_corte = date.today()

    alumnos = extraer_alumnos_pdf(ruta_pdf)
    if not alumnos:
        raise ValueError("No se encontraron alumnos en el PDF.")

    nombre, direccion, codigo = extraer_metadatos_encabezado_pdf(ruta_pdf)
    nombre_escuela = construir_nombre_escolar_completo(nombre, direccion)
    if not nombre_escuela:
        nombre_escuela = "ESCUELA SIN NOMBRE"

    # Convertir todos los alumnos y filtrar aptos por edad
    todos_formateados = []
    aptos = []
    for a in alumnos:
        dia, mes, anio = _split_fecha(a["fecha_nac"])
        gen = "F" if a.get("genero", "").startswith("F") else "M"
        nombre_completo_alumno = _nombre_completo(a)
        cui = a.get("cui", "")
        fecha_nac = a.get("fecha_nac", "")

        todos_formateados.append({
            "nombre": nombre_completo_alumno,
            "cui": cui,
            "genero": gen,
            "dia": dia,
            "mes": mes,
            "anio": anio,
            "fecha_nac": fecha_nac,
            "grado": a.get("grado", ""),
            "seccion": a.get("seccion", ""),
        })

        edad = calcular_edad_a_fecha_corte(dia, mes, anio, fecha_corte)
        if edad is not None and 6 <= edad <= 14:
            aptos.append({
                "nombre": nombre_completo_alumno,
                "cui": cui,
                "genero": gen,
                "dia": dia,
                "mes": mes,
                "anio": anio,
                "grado": a.get("grado", ""),
                "seccion": a.get("seccion", ""),
            })

    os.makedirs(ruta_salida_dir, exist_ok=True)
    nombre_archivo = f"SISCA_{_nombre_archivo_seguro(codigo or nombre_escuela)}.xlsx"
    ruta_salida = os.path.join(ruta_salida_dir, nombre_archivo)

    anio_campana = fecha_corte.year
    generar_ficha_sisca_escuela(
        ruta_plantilla, ruta_salida,
        nombre_escuela, codigo, aptos, anio_campana,
        tipo_centro=tipo_centro,
        responsable=responsable, cargo=cargo,
        area=area_salud, distrito=distrito_salud,
        servicio=servicio_salud,
        fecha_reporte_str=fecha_reporte_str,
    )

    return {
        "ruta": ruta_salida,
        "escuela_nombre": nombre_escuela,
        "escuela_codigo": codigo,
        "tipo_centro": tipo_centro,
        "servicio_salud": servicio_salud,
        "alumnos": todos_formateados,
        "total_aptos": len(aptos),
    }


# =============================================================================
#  EXPORTACIÓN — CONSOLIDADO E HISTORIAL
# =============================================================================

_STYLE_HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_STYLE_DATA   = Font(name="Arial", size=10)
_FILL_HEADER  = PatternFill("solid", fgColor="0C6478")
_FILL_ALT     = PatternFill("solid", fgColor="E8EDF5")


GRADO_ORDEN = {
    "PARVULOS": 0, "PREPRIMARIA": 0, "PREPRIMARIO": 0,
    "PRIMERO": 1, "PRIMER": 1,
    "SEGUNDO": 2, "SEGUND": 2,
    "TERCERO": 3, "TERCER": 3,
    "CUARTO": 4,
    "QUINTO": 5,
    "SEXTO": 6,
}

def _orden_grado(g):
    """Retorna (base_order, sub_level) para ordenar por grado con sufijo numérico."""
    g_clean = _quitar_acentos(g.strip().upper())
    # Extraer sufijo numérico (ej: "PARVULOS 3" → base="PARVULOS", sub=3)
    m = re.search(r'^([A-Z\s]+?)\s*(\d+)$', g_clean)
    if m:
        base = m.group(1).strip()
        sub = int(m.group(2))
    else:
        base = g_clean
        sub = 0
    orden = GRADO_ORDEN.get(base, 99)
    return (orden, sub)

def _clave_grado_seccion(grado, seccion):
    """Sort key para tuplas (grado, seccion)."""
    orden, sub = _orden_grado(grado)
    return (orden, sub, seccion or "")

def _clasificar_rango(edad):
    if edad is None:
        return None
    if edad <= 5:
        return "5_y_menos"
    if edad <= 9:
        return "6_a_9"
    if edad <= 14:
        return "10_a_14"
    return "15_a_19"  # 15 años o más → rango único


def _edad_desde_fecha_nac(fecha_nac_str, fecha_corte: date):
    try:
        partes = str(fecha_nac_str).split("/")
        dia, mes, anio = int(partes[0]), int(partes[1]), int(partes[2])
        edad = fecha_corte.year - anio
        if (fecha_corte.month, fecha_corte.day) < (mes, dia):
            edad -= 1
        return edad
    except Exception:
        return None


def generar_excel_escuela(ruta_salida: str,
                          escuela_nombre: str, codigo: str,
                          alumnos: list, fecha_corte: date):
    """Libro individual: listado nominal detallado de una escuela, agrupado por Grado/Sección."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado Nominal"

    fuente_meta = Font(bold=True, size=11, name="Arial")
    fuente_grupo = Font(bold=True, size=12, name="Arial", color="FFFFFF")
    fill_grupo = PatternFill("solid", fgColor="0C6478")
    fill_alerta = PatternFill("solid", fgColor="FFC7CE")
    fuente_alerta = Font(name="Arial", size=10, color="9C0006")

    # Encabezados fijos
    fila = 1
    for celda, val in [("A1", "Fecha de corte:"), ("C1", fecha_corte.strftime("%d/%m/%Y")),
                        ("A2", "Nombre de la Escuela:"), ("C2", escuela_nombre),
                        ("A3", "Código:"), ("C3", codigo)]:
        ws[celda] = val
        ws[celda].font = fuente_meta

    # Columnas del listado
    headers = ["No.", "Nombre", "CUI", "Género", "Día", "Mes", "Año",
               "Edad", "<=5", "6-9", "10-14", "15-19"]
    ancho_cols = [6, 38, 16, 8, 6, 6, 8, 6, 6, 6, 8, 8]
    for i, w in enumerate(ancho_cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Agrupar por (grado, seccion) y ordenar
    grupos = {}
    for a in alumnos:
        key = (a.get("grado", "") or "", a.get("seccion", "") or "")
        grupos.setdefault(key, []).append(a)
    for k in grupos:
        grupos[k].sort(key=lambda x: x.get("nombre", ""))

    fila = 5

    for (grado, seccion), grupo in sorted(grupos.items(),
                                          key=lambda item: _clave_grado_seccion(item[0][0], item[0][1])):
        # Fila de subtítulo del grupo (ocupando todo el ancho)
        fila += 1
        texto_grupo = f"{grado or 'SIN GRADO'} — Sección {seccion or 'SIN SECCIÓN'}"
        ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=12)
        cel_sub = ws.cell(fila, 1, texto_grupo)
        cel_sub.font = fuente_grupo
        cel_sub.fill = fill_grupo
        cel_sub.alignment = Alignment(horizontal="left", vertical="center")
        for c in range(1, 13):
            ws.cell(fila, c).fill = fill_grupo

        # Encabezados de columna
        fila += 1
        for col, h in enumerate(headers, 1):
            c = ws.cell(fila, col, h)
            c.font = _STYLE_HEADER
            c.fill = _FILL_HEADER
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Alumnos del grupo (correlativo reinicia en 1)
        for i_local, a in enumerate(grupo, start=1):
            fila += 1
            edad = _edad_desde_fecha_nac(a.get("fecha_nac", ""), fecha_corte)
            rango = _clasificar_rango(edad)

            ws.cell(fila, 1, i_local).font = _STYLE_DATA
            ws.cell(fila, 2, a.get("nombre", "")).font = _STYLE_DATA
            cel_cui = ws.cell(fila, 3, str(a.get("cui", "")))
            cel_cui.data_type = "s"
            cel_cui.font = _STYLE_DATA
            ws.cell(fila, 4, a.get("genero", "")).font = _STYLE_DATA
            try:
                dia, mes, anio = str(a.get("fecha_nac", "")).split("/")
                ws.cell(fila, 5, int(dia)).font = _STYLE_DATA
                ws.cell(fila, 6, int(mes)).font = _STYLE_DATA
                ws.cell(fila, 7, int(anio)).font = _STYLE_DATA
            except Exception:
                pass

            cel_edad = ws.cell(fila, 8, edad if edad is not None else "")
            cel_edad.font = _STYLE_DATA

            for rcol, rname in [(9, "5_y_menos"), (10, "6_a_9"), (11, "10_a_14"), (12, "15_a_19")]:
                ws.cell(fila, rcol, "X" if rango == rname else "").font = _STYLE_DATA

            fuera_rango = edad is not None and (edad < 6 or edad > 14)
            for col in range(1, 13):
                cel = ws.cell(fila, col)
                cel.alignment = Alignment(horizontal="center", vertical="center")
                if fuera_rango:
                    cel.fill = fill_alerta
                    cel.font = fuente_alerta
                elif i_local % 2 == 0:
                    cel.fill = _FILL_ALT

    wb.save(ruta_salida)
    return ruta_salida


def generar_excel_consolidado(ruta_salida: str,
                              matriz: list,
                              escuelas_detalle: dict,
                              fecha_corte: date):
    """
    Libro completo: Hoja1 = matriz general, hojas siguientes = cada escuela.
    `matriz`: lista de dicts con claves nombre, codigo y conteos por rango+genero.
    `escuelas_detalle`: dict {codigo: [alumnos]}.
    """
    wb = openpyxl.Workbook()

    # ── Hoja 1: Matriz consolidada ────────────────────────────────────────
    ws = wb.active
    ws.title = "Consolidado"

    ws["A1"] = "CONSOLIDADO DE SALUD RURAL"
    ws["A1"].font = Font(bold=True, size=14, name="Arial", color="0C6478")
    ws.merge_cells("A1:O1")
    ws["A2"] = f"Corte: {fecha_corte.strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(size=10, name="Arial", color="555555")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ancho_cols = [5, 38, 18, 7, 7, 7, 7, 8, 8, 8, 8, 9, 9, 13, 11]
    for i, w in enumerate(ancho_cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Encabezados combinados (Fila 4: bloques, Fila 5: M/F) ----
    merge_bloques = [
        ("A4:A5", "No."), ("B4:B5", "Establecimiento"),
        ("C4:C5", "Código Centro"),
        ("D4:E4", "Inscritos"), ("F4:G4", "≤ 5 años"), ("H4:I4", "6 - 9 años"),
        ("J4:K4", "10 - 14 años"), ("L4:M4", "15 - 19 años"),
        ("N4:N5", "Total Escolarizados\n(6 a 14)"), ("O4:O5", "Total General"),
    ]

    def _estilar_encabezado(celda):
        celda.font = _STYLE_HEADER
        celda.fill = _FILL_HEADER
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = thin_border

    for rango, texto in merge_bloques:
        ws.merge_cells(rango)
        cel = ws[rango.split(":")[0]]
        cel.value = texto
        _estilar_encabezado(cel)

    # Sub-etiquetas M/F en fila 5 (solo columnas no combinadas verticalmente)
    ws.cell(5, 4, "M"); ws.cell(5, 5, "F")   # Inscritos
    ws.cell(5, 6, "M"); ws.cell(5, 7, "F")   # ≤ 5 años
    ws.cell(5, 8, "M"); ws.cell(5, 9, "F")   # 6 - 9 años
    ws.cell(5, 10, "M"); ws.cell(5, 11, "F")  # 10 - 14 años
    ws.cell(5, 12, "M"); ws.cell(5, 13, "F")  # 15 - 19 años

    fila = 5
    # Una fila por establecimiento (suma de todos sus grados)
    no_global = 0
    for codigo, alumnos in escuelas_detalle.items():
        nombre_escuela = next((m["nombre"] for m in matriz if m["codigo"] == codigo), codigo)

        m_tot = f_tot = m_5 = f_5 = m_6_9 = f_6_9 = m_10_14 = f_10_14 = m_15_19 = f_15_19 = 0
        for a in alumnos:
            gen = a.get("genero", "").upper()
            if gen == "M":
                m_tot += 1
            else:
                f_tot += 1
            edad = a.get("edad")
            if edad is not None:
                if edad <= 5:
                    if gen == "M": m_5 += 1
                    else: f_5 += 1
                elif edad <= 9:
                    if gen == "M": m_6_9 += 1
                    else: f_6_9 += 1
                elif edad <= 14:
                    if gen == "M": m_10_14 += 1
                    else: f_10_14 += 1
                else:
                    if gen == "M": m_15_19 += 1
                    else: f_15_19 += 1

        no_global += 1
        fila += 1
        total_esc = m_6_9 + f_6_9 + m_10_14 + f_10_14
        total_gen = m_tot + f_tot

        ws.cell(fila, 1, no_global).font = _STYLE_DATA
        ws.cell(fila, 2, nombre_escuela).font = _STYLE_DATA
        ws.cell(fila, 3, codigo).font = _STYLE_DATA

        vals = [m_tot, f_tot, m_5, f_5, m_6_9, f_6_9, m_10_14, f_10_14, m_15_19, f_15_19, total_esc, total_gen]
        for col_idx, val in enumerate(vals, start=4):
            c = ws.cell(fila, col_idx, val)
            c.font = _STYLE_DATA
            c.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(fila, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(fila, 3).alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 16):
            cel = ws.cell(fila, col)
            cel.border = thin_border
            if no_global % 2 == 0:
                cel.fill = _FILL_ALT

    # ── Fila de TOTALES al final de la matriz ──────────────────────────────
    total_row = fila + 1
    fuente_total = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_total = PatternFill("solid", fgColor="213A58")
    ws.cell(total_row, 1).font = fuente_total
    ws.cell(total_row, 1).fill = fill_total
    ws.cell(total_row, 1).border = thin_border
    ws.cell(total_row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(total_row, 2, "TOTALES").font = fuente_total
    ws.cell(total_row, 2).fill = fill_total
    ws.cell(total_row, 2).border = thin_border
    ws.cell(total_row, 3).fill = fill_total
    ws.cell(total_row, 3).border = thin_border
    first_data = 6
    for col in range(4, 16):
        cel = ws.cell(total_row, col)
        cel.value = f"=SUM({openpyxl.utils.get_column_letter(col)}{first_data}:{openpyxl.utils.get_column_letter(col)}{total_row - 1})"
        cel.font = fuente_total
        cel.fill = fill_total
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = thin_border

    # ── Hojas siguientes: cada escuela (agrupadas por grado/sección) ──────
    fuente_grupo_cons = Font(bold=True, size=12, name="Arial", color="FFFFFF")
    fill_grupo_cons = PatternFill("solid", fgColor="0C6478")
    fill_alerta_cons = PatternFill("solid", fgColor="FFC7CE")
    fuente_alerta_cons = Font(name="Arial", size=10, color="9C0006")

    for idx, (codigo, alumnos) in enumerate(escuelas_detalle.items()):
        if idx == 0:
            ws_det = wb.create_sheet(title=f"ESC {codigo[-4:]}" if len(codigo) > 4 else codigo, index=1)
        else:
            ws_det = wb.create_sheet(title=f"ESC {codigo[-4:]}" if len(codigo) > 4 else codigo)

        ws_det["A1"] = "Listado Nominal"
        ws_det["A1"].font = Font(bold=True, size=12, name="Arial", color="0C6478")
        ws_det["A2"] = f"Código: {codigo}"
        ws_det["A2"].font = Font(size=10, name="Arial")

        det_headers = ["No.", "Nombre", "CUI", "Género",
                       "Día", "Mes", "Año", "Edad", "<=5", "6-9", "10-14", "15-19"]
        det_ancho = [6, 38, 16, 8, 6, 6, 8, 6, 6, 6, 8, 8]
        for i, w in enumerate(det_ancho, 1):
            ws_det.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        # Agrupar alumnos por (grado, seccion)
        grupos_det = {}
        for a in alumnos:
            key = (a.get("grado", "") or "", a.get("seccion", "") or "")
            grupos_det.setdefault(key, []).append(a)
        for k in grupos_det:
            grupos_det[k].sort(key=lambda x: x.get("nombre", ""))

        fila = 4

        for (grado, seccion), grupo in sorted(grupos_det.items(),
                                              key=lambda item: _clave_grado_seccion(item[0][0], item[0][1])):
            # Fila de subtítulo del grupo
            fila += 1
            texto_grupo = f"{grado or 'SIN GRADO'} — Sección {seccion or 'SIN SECCIÓN'}"
            ws_det.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=12)
            cel_sub = ws_det.cell(fila, 1, texto_grupo)
            cel_sub.font = fuente_grupo_cons
            cel_sub.fill = fill_grupo_cons
            cel_sub.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 13):
                ws_det.cell(fila, c).fill = fill_grupo_cons

            # Encabezados de columna
            fila += 1
            for col, h in enumerate(det_headers, 1):
                c = ws_det.cell(fila, col, h)
                c.font = _STYLE_HEADER
                c.fill = _FILL_HEADER
                c.alignment = Alignment(horizontal="center")

            # Alumnos del grupo (correlativo reinicia en 1)
            for i_local, a in enumerate(grupo, start=1):
                fila += 1
                edad = _edad_desde_fecha_nac(a.get("fecha_nac", ""), fecha_corte)
                rango = _clasificar_rango(edad)
                ws_det.cell(fila, 1, i_local).font = _STYLE_DATA
                ws_det.cell(fila, 2, a.get("nombre", "")).font = _STYLE_DATA
                cel_cui = ws_det.cell(fila, 3, str(a.get("cui", "")))
                cel_cui.data_type = "s"
                cel_cui.font = _STYLE_DATA
                ws_det.cell(fila, 4, a.get("genero", "")).font = _STYLE_DATA
                try:
                    d, m, y = str(a.get("fecha_nac", "")).split("/")
                    ws_det.cell(fila, 5, int(d)).font = _STYLE_DATA
                    ws_det.cell(fila, 6, int(m)).font = _STYLE_DATA
                    ws_det.cell(fila, 7, int(y)).font = _STYLE_DATA
                except Exception:
                    pass
                cel_edad = ws_det.cell(fila, 8, edad if edad is not None else "")
                cel_edad.font = _STYLE_DATA
                for rcol, rname in [(9, "5_y_menos"), (10, "6_a_9"), (11, "10_a_14"), (12, "15_a_19")]:
                    ws_det.cell(fila, rcol, "X" if rango == rname else "").font = _STYLE_DATA
                fuera_rango = edad is not None and (edad < 6 or edad > 14)
                for c in range(1, 13):
                    ws_det.cell(fila, c).alignment = Alignment(horizontal="center")
                    if fuera_rango:
                        ws_det.cell(fila, c).fill = fill_alerta_cons
                        ws_det.cell(fila, c).font = fuente_alerta_cons

    wb.save(ruta_salida)
    return ruta_salida


# =============================================================================
#  GENERACIÓN — SIGSA-22 (Fluorización)
# =============================================================================

_STUDENT_ROWS_START = 17
def generar_reporte_sigsa22(ruta_plantilla, ruta_salida, datos):
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(f"No se encontró la plantilla SIGSA-22 en:\n{ruta_plantilla}")

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws_template = wb.active
    MAX_FILAS = 15

    estudiantes = datos.get("estudiantes", [])
    total_paginas = max(1, math.ceil(len(estudiantes) / MAX_FILAS))

    # ── Crear todas las páginas (clonando desde la plantilla virgen) ────────
    paginas = []
    for p in range(total_paginas):
        if p == 0:
            ws = ws_template
        else:
            _asegurar_sheet_properties(ws_template)
            ws = wb.copy_worksheet(ws_template)
        ws.title = f"SIGSA-22 Pág {p + 1}"

        try:
            ws.print_area = ws_template.print_area
            if hasattr(ws_template, "page_setup") and ws_template.page_setup:
                ws.page_setup.orientation = ws_template.page_setup.orientation
                ws.page_setup.paperSize = ws_template.page_setup.paperSize
                ws.page_setup.fitToPage = ws_template.page_setup.fitToPage
        except Exception:
            pass

        paginas.append(ws)

    # ── Inyectar datos en cada página ───────────────────────────────────────
    fuente_premium = Font(name="Arial", size=14, bold=True, color="FF000000")

    for idx, ws in enumerate(paginas):
        # ── Encabezado ──────────────────────────────────────────────────────────
        ws["C7"].value = datos.get("area_salud", "")
        ws["C7"].font = fuente_premium
        ws["T7"].value = datos.get("distrito_salud", "")
        ws["T7"].font = fuente_premium
        ws["AK7"].value = datos.get("municipio", "")
        ws["AK7"].font = fuente_premium
        ws["AZ7"].value = datos.get("servicio_salud", "")
        ws["AZ7"].font = fuente_premium
        ws["F9"].value = datos.get("responsable_informacion", "")
        ws["F9"].font = fuente_premium
        ws["R9"].value = datos.get("cargo", "")
        ws["R9"].font = fuente_premium
        ws["BE9"].value = datos.get("mes_reporte", "")
        ws["BE9"].font = fuente_premium
        ws["BN9"].value = datos.get("anio_reporte", "")
        ws["BN9"].font = fuente_premium

        # ── Estudiantes de este bloque (fila 17 a 31) ───────────────────────
        inicio = idx * MAX_FILAS
        bloque = estudiantes[inicio:inicio + MAX_FILAS]

        for j, estudiante in enumerate(bloque):
            fila_actual = 17 + j
            idx_global = inicio + j

            ws.cell(row=fila_actual, column=1, value=idx_global + 1).font = fuente_premium
            ws.cell(row=fila_actual, column=3, value=str(estudiante.get("nombre_completo", "")).strip()).font = fuente_premium

            celda_cui = ws.cell(row=fila_actual, column=13)
            celda_cui.data_type = "s"
            celda_cui.value = str(estudiante.get("cui", "")).strip()
            celda_cui.font = fuente_premium

            ws.cell(row=fila_actual, column=14, value=str(estudiante.get("sexo", "")).strip()).font = fuente_premium
            ws.cell(row=fila_actual, column=15, value=str(estudiante.get("pueblo", "")).strip()).font = fuente_premium
            ws.cell(row=fila_actual, column=16, value=str(estudiante.get("comunidad_linguistica", "")).strip()).font = fuente_premium
            ws.cell(row=fila_actual, column=17, value=estudiante.get("dia_nac", "")).font = fuente_premium
            ws.cell(row=fila_actual, column=18, value=estudiante.get("mes_nac", "")).font = fuente_premium
            ws.cell(row=fila_actual, column=19, value=estudiante.get("anio_nac", "")).font = fuente_premium

    wb.save(ruta_salida)
    wb.close()
    gc.collect()
