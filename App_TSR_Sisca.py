"""
TSR - Herramienta de Salud Rural  (App_TSR_Sisca.py)
=====================================================
Versión PyQt6 con interfaz Glassmorphism de transparencia real.
"""

import os, re, sys, math, copy, gc, unicodedata, json
from datetime import date, datetime
from io import BytesIO

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QLabel,
    QFrame, QTabWidget, QPushButton, QComboBox, QLineEdit, QFileDialog,
    QMessageBox, QInputDialog, QDialog, QDialogButtonBox,
    QVBoxLayout, QHBoxLayout, QFormLayout, QSizePolicy)
from PyQt6.QtCore import Qt, QRect, QTimer, QSize
from PyQt6.QtGui import (QFont, QPixmap, QPainter, QLinearGradient, QColor,
    QBrush, QPen)

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage, _import_image


# =============================================================================
#  CONFIGURACIÓN INSTITUCIONAL
# =============================================================================

RESPONSABLE_NOMBRE  = ""
RESPONSABLE_CARGO   = ""
DISTRITO_DE_SALUD   = ""
AREA_DE_SALUD       = ""
SERVICIO_DE_SALUD   = ""
TIPO_CENTRO_EDUCATIVO_DEFECTO = "PUBLICO"

ANIO_CAMPANA_DEFECTO = 2026
FECHA_CORTE_DEFECTO  = (31, 3, 2026)

NOMBRES_ARCHIVOS = {
    "base_datos":      "Base_Datos_Consolidado_TSR.xlsx",
    "plantilla_sisca":  "plantilla_sisca.xlsx",
    "carpeta_sisca":   "Fichas_SISCA_generadas",
    "fondo":           "fondo_tsr.jpg",
    "icono":           "logo_tsr.ico",
}


# =============================================================================
#  RUTAS DE ARCHIVOS
# =============================================================================

def _ruta_junto_al_ejecutable(nombre_archivo: str) -> str:
    if getattr(sys, "frozen", False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, nombre_archivo)


def _obtener_ruta_recurso(nombre_archivo: str) -> str:
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, nombre_archivo)


def ruta_base_datos() -> str:
    return _ruta_junto_al_ejecutable(NOMBRES_ARCHIVOS["base_datos"])


def ruta_plantilla_sisca() -> str:
    return _ruta_junto_al_ejecutable(NOMBRES_ARCHIVOS["plantilla_sisca"])


def carpeta_salida_sisca() -> str:
    carpeta = _ruta_junto_al_ejecutable(NOMBRES_ARCHIVOS["carpeta_sisca"])
    os.makedirs(carpeta, exist_ok=True)
    return carpeta


# =============================================================================
#  CONFIGURACIÓN LOCAL PERSISTENTE (config_tsr.json)
# =============================================================================

CONFIG_TSR_FILE = "config_tsr.json"

def _ruta_config_tsr() -> str:
    return _ruta_junto_al_ejecutable(CONFIG_TSR_FILE)

def cargar_config_tsr() -> dict:
    ruta = _ruta_config_tsr()
    if os.path.exists(ruta):
        try:
            with open(ruta, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {
        "responsable": "",
        "cargo": "",
        "area_salud": "",
        "distrito_salud": "",
        "servicio_salud": "",
        "tipo_centro": "PUBLICO",
    }

def guardar_config_tsr(datos: dict):
    ruta = _ruta_config_tsr()
    try:
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(datos, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


# =============================================================================
#  UTILIDADES DE TEXTO / NÚMEROS
# =============================================================================

def _quitar_acentos(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

def _normalizar(texto: str) -> str:
    return _quitar_acentos(str(texto or "")).strip().lower()

def _nombre_archivo_seguro(texto: str) -> str:
    limpio = _quitar_acentos(str(texto or "SIN_NOMBRE"))
    limpio = re.sub(r'[\\/*?:"<>|]', "", limpio)
    limpio = re.sub(r"\s+", "_", limpio.strip())
    return limpio[:80] if limpio else "ESCUELA"

def _split_fecha(fecha: str):
    try:
        p = str(fecha).split("/")
        return int(p[0]), int(p[1]), int(p[2])
    except Exception:
        return None, None, None

def _nombre_completo(alumno: dict) -> str:
    return f"{alumno['nombres']} {alumno['apellidos']}"

def calcular_edad(anio_nacimiento, anio_campana: int):
    try:
        return int(anio_campana) - int(anio_nacimiento)
    except Exception:
        return None

def calcular_edad_a_fecha_corte(dia_nac, mes_nac, anio_nac, fecha_corte: date):
    try:
        dia_nac, mes_nac, anio_nac = int(dia_nac), int(mes_nac), int(anio_nac)
        edad = fecha_corte.year - anio_nac
        if (fecha_corte.month, fecha_corte.day) < (mes_nac, dia_nac):
            edad -= 1
        return edad
    except Exception:
        return None


# =============================================================================
#  EXTRACCIÓN DE ALUMNOS DEL PDF
# =============================================================================

def extraer_alumnos_pdf(ruta_pdf: str) -> list:
    alumnos = []
    grado_actual   = ""
    seccion_actual = ""

    def es_fila_alumno(fila):
        return (fila and len(fila) >= 9 and str(fila[0] or "").strip().isdigit())

    def procesar_fila(fila):
        ap  = str(fila[2] or "").replace("\n", " ").strip().upper()
        nom = str(fila[3] or "").replace("\n", " ").strip().upper()
        fec = str(fila[4] or "").strip()
        cui = str(fila[7] or "").strip()
        gen = str(fila[8] or "").strip().upper()
        if ap and nom:
            alumnos.append({
                "grado":     grado_actual,
                "seccion":   seccion_actual,
                "apellidos": ap,
                "nombres":   nom,
                "fecha_nac": fec,
                "cui":       cui,
                "genero":    gen,
            })

    with pdfplumber.open(ruta_pdf) as pdf:
        for pagina in pdf.pages:
            for tabla in pagina.extract_tables():
                if not tabla:
                    continue
                enc = [str(c).replace("\n", " ").strip() if c else "" for c in tabla[0]]

                if enc[0] == "Grado:" and len(enc) >= 4:
                    grado_actual   = str(enc[1] or "").strip().upper()
                    seccion_actual = str(enc[3] or "").strip().upper()
                    continue

                if "Apellidos" in enc and "Nombres" in enc:
                    for fila in tabla[1:]:
                        if es_fila_alumno(fila):
                            get_row = fila
                            procesar_fila(get_row)
                    continue

                if enc[0].isdigit() and len(enc) >= 9:
                    for fila in tabla:
                        if es_fila_alumno(fila):
                            procesar_fila(fila)

    return alumnos


# =============================================================================
#  EXTRACCIÓN DE METADATOS DEL ENCABEZADO
# =============================================================================

_ETIQUETAS_ENCABEZADO_PDF = ["Nombre:", "Dirección:", "Código:"]
_MAX_LINEAS_ENCABEZADO = 15

def _valor_tras_etiqueta(etiqueta: str, texto: str) -> str:
    idx = texto.find(etiqueta)
    if idx == -1:
        return ""
    resto = texto[idx + len(etiqueta):]

    cortes = [resto.find(e) for e in _ETIQUETAS_ENCABEZADO_PDF if e != etiqueta and resto.find(e) != -1]
    idx_salto = resto.find("\n")
    if idx_salto != -1:
        cortes.append(idx_salto)
    fin = min(cortes) if cortes else len(resto)

    valor = resto[:fin].replace("\n", " ")
    valor = re.sub(r"\s+", " ", valor).strip()
    return valor

def extraer_metadatos_encabezado_pdf(ruta_pdf: str):
    try:
        with pdfplumber.open(ruta_pdf) as pdf:
            texto_pag1 = pdf.pages[0].extract_text() or ""
    except Exception:
        return "", "", ""

    lineas = texto_pag1.splitlines()[:_MAX_LINEAS_ENCABEZADO]
    texto_top = "\n".join(lineas)

    nombre    = _valor_tras_etiqueta("Nombre:", texto_top).upper()
    direccion = _valor_tras_etiqueta("Dirección:", texto_top).upper()
    codigo_raw = _valor_tras_etiqueta("Código:", texto_top).upper()

    m = re.search(r"\d{1,2}-\d{1,2}-\d{2,5}-\d{1,3}", codigo_raw)
    codigo = m.group(0) if m else codigo_raw

    return nombre, direccion, codigo

def construir_nombre_escolar_completo(nombre: str, direccion: str) -> str:
    combinado = f"{nombre.strip()} {direccion.strip()}".strip()
    combinado = combinado.replace("\n", " ")
    combinado = re.sub(r"\s+", " ", combinado)
    return combinado.upper()


# =============================================================================
#  BASE DE DATOS CENTRAL
# =============================================================================

FILA_INICIO_DATOS_ESCUELA = 6
FILA_TOTAL_ESCUELA        = 501
TOTAL_HOJAS_ESCUELA       = 40
COLOR_SEPARADOR_ESCUELA = "3B48F6"

def abrir_base_datos() -> openpyxl.Workbook:
    ruta = ruta_base_datos()
    if not os.path.exists(ruta):
        raise FileNotFoundError(
            f"No se encontró la Base de Datos Central en:\n{ruta}\n\n"
            "Coloca el archivo 'Base_Datos_Consolidado_TSR.xlsx' junto al programa."
        )
    return openpyxl.load_workbook(ruta)

def _hoja_escuela(wb, numero: int):
    return wb[f"ESCUELA {numero}"]

def _hoja_esta_vacia(ws) -> bool:
    for fila in range(FILA_INICIO_DATOS_ESCUELA, FILA_TOTAL_ESCUELA):
        if str(ws.cell(fila, 3).value or "").strip():
            return False
    return True

def _hoja_disponible_para_nueva_escuela(ws) -> bool:
    m2 = str(ws["M2"].value or "").strip().upper()
    return m2 == "" or m2 == "EORM"

def _ultima_fila_libre(ws) -> int:
    for fila in range(FILA_INICIO_DATOS_ESCUELA, FILA_TOTAL_ESCUELA):
        b_vacio = str(ws.cell(fila, 2).value or "").strip() == ""
        c_vacio = str(ws.cell(fila, 3).value or "").strip() == ""
        if b_vacio and c_vacio:
            return fila
    return FILA_TOTAL_ESCUELA

def encontrar_hoja_para_escuela(wb, codigo: str, nombre_completo: str):
    codigo_n = _normalizar(codigo)
    nombre_n = _normalizar(nombre_completo)

    if codigo_n:
        for i in range(1, TOTAL_HOJAS_ESCUELA + 1):
            ws = _hoja_escuela(wb, i)
            if _hoja_disponible_para_nueva_escuela(ws):
                continue
            if _normalizar(ws["P2"].value) == codigo_n:
                return ws

    if nombre_n:
        for i in range(1, TOTAL_HOJAS_ESCUELA + 1):
            ws = _hoja_escuela(wb, i)
            if _hoja_disponible_para_nueva_escuela(ws):
                continue
            if _normalizar(ws["M2"].value) == nombre_n:
                return ws
    return None

def encontrar_hoja_libre(wb):
    for i in range(1, TOTAL_HOJAS_ESCUELA + 1):
        ws = _hoja_escuela(wb, i)
        if _hoja_disponible_para_nueva_escuela(ws):
            return ws
    return None

def _borde_fino():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _fijar_encabezado_escuela(ws, nombre_completo: str, codigo: str, fecha_corte: date):
    ws["M2"] = nombre_completo.strip().upper()
    ws["P2"] = codigo.strip().upper()
    try:
        ws["C2"] = fecha_corte
    except Exception:
        pass

def _escribir_separador_grado(ws, fila: int, texto: str):
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == fila and rng.max_row == fila and rng.min_col <= 12:
            ws.unmerge_cells(str(rng))

    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=12)
    celda = ws.cell(fila, 2, texto)
    celda.font = Font(bold=True, size=14, name="Arial", color="FFFFFF")
    celda.fill = PatternFill("solid", start_color=COLOR_SEPARADOR_ESCUELA)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 16.55

def _estilo_celda_dato(ws, fila, col, valor, rotacion=0, wrap=False):
    c = ws.cell(fila, col, valor)
    c.font = Font(bold=True, size=14, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center", text_rotation=rotacion, wrap_text=wrap)
    c.border = _borde_fino()
    return c

def _escribir_fila_alumno(ws, fila: int, correlativo: int, alumno: dict):
    dia, mes, anio = _split_fecha(alumno["fecha_nac"])
    nombre = _nombre_completo(alumno)
    es_femenino  = alumno.get("genero", "").startswith("F")
    es_masculino = alumno.get("genero", "").startswith("M")

    _estilo_celda_dato(ws, fila, 2, correlativo)

    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == fila and rng.max_row == fila and 2 <= rng.min_col <= 18 and (rng.max_col - rng.min_col + 1) != 10:
            ws.unmerge_cells(str(rng))
    ya_combinada = any(
        rng.min_row == fila and rng.max_row == fila and rng.min_col == 3 and rng.max_col == 12
        for rng in ws.merged_cells.ranges
    )
    if not ya_combinada:
        ws.merge_cells(start_row=fila, start_column=3, end_row=fila, end_column=12)

    _estilo_celda_dato(ws, fila, 3, nombre, wrap=True)

    celda_cui = _estilo_celda_dato(ws, fila, 13, alumno.get("cui", ""))
    celda_cui.number_format = "@"

    _estilo_celda_dato(ws, fila, 14, "X" if es_femenino else "")
    _estilo_celda_dato(ws, fila, 15, "X" if es_masculino else "")
    _estilo_celda_dato(ws, fila, 16, dia)
    _estilo_celda_dato(ws, fila, 17, mes)
    _estilo_celda_dato(ws, fila, 18, anio, rotacion=90)

    ws.row_dimensions[fila].height = 16.55

def agregar_alumnos_a_hoja(ws, alumnos: list) -> int:
    fila = _ultima_fila_libre(ws)
    grado_ant = seccion_ant = object()
    correlativo = 1
    agregados = 0

    for alumno in alumnos:
        if fila >= FILA_TOTAL_ESCUELA:
            break

        g, s = alumno.get("grado", ""), alumno.get("seccion", "")
        if g != grado_ant or s != seccion_ant:
            if fila >= FILA_TOTAL_ESCUELA:
                break
            etiqueta = f"--- {g} - SECCIÓN {s} ---" if g or s else "--- SIN GRADO ---"
            _escribir_separador_grado(ws, fila, etiqueta)
            fila += 1
            correlativo = 1
            grado_ant, seccion_ant = g, s

        if fila >= FILA_TOTAL_ESCUELA:
            break

        _escribir_fila_alumno(ws, fila, correlativo, alumno)
        fila += 1
        correlativo += 1
        agregados += 1

    return agregados


# =============================================================================
#  FLUJO INTERNACIÓN DE ESCUELAS DESDE PDF
# =============================================================================

def insertar_nueva_escuela_desde_pdf(ruta_pdf: str, fecha_corte: date, pedir_dato_manual=None) -> dict:
    alumnos = extraer_alumnos_pdf(ruta_pdf)
    if not alumnos:
        raise ValueError("No se encontraron alumnos en el listado nominal del PDF.")

    nombre, direccion, codigo = extraer_metadatos_encabezado_pdf(ruta_pdf)

    if not nombre and pedir_dato_manual:
        nombre = pedir_dato_manual("No se detectó 'Nombre:' en el encabezado del PDF.\nEscribe el nombre de la escuela:") or ""
    if not direccion and pedir_dato_manual:
        direccion = pedir_dato_manual("No se detectó 'Dirección:' en el encabezado del PDF.\nEscribe la dirección:") or ""
    if not codigo and pedir_dato_manual:
        codigo = pedir_dato_manual("No se detectó 'Código:' en el encabezado del PDF.\nEscribe el código:") or ""

    nombre_completo = construir_nombre_escolar_completo(nombre, direccion)
    if not nombre_completo and not codigo:
        raise ValueError("Falta nombre/dirección y código de la escuela; no se puede continuar.")

    wb = abrir_base_datos()
    ws = encontrar_hoja_para_escuela(wb, codigo, nombre_completo)
    es_nueva = ws is None
    if es_nueva:
        ws = encontrar_hoja_libre(wb)
        if ws is None:
            raise RuntimeError(f"Ya existen {TOTAL_HOJAS_ESCUELA} escuelas registradas; no hay hojas disponibles.")
        _fijar_encabezado_escuela(ws, nombre_completo, codigo, fecha_corte)
    else:
        ws["C2"] = fecha_corte

    agregados = agregar_alumnos_a_hoja(ws, alumnos)

    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass
    wb.save(ruta_base_datos())
    wb.close()
    gc.collect()

    return {
        "hoja": ws.title,
        "nombre_completo": nombre_completo,
        "codigo": codigo,
        "agregados": agregados,
        "total_alumnos_pdf": len(alumnos),
        "es_nueva": es_nueva,
    }

def listar_escuelas_procesadas() -> list:
    if not os.path.exists(ruta_base_datos()):
        return []
    wb = openpyxl.load_workbook(ruta_base_datos(), read_only=True, data_only=True)
    escuelas = []
    for i in range(1, TOTAL_HOJAS_ESCUELA + 1):
        titulo = f"ESCUELA {i}"
        if titulo not in wb.sheetnames:
            continue
        ws = wb[titulo]
        m2 = str(ws["M2"].value or "").strip()
        if not m2 or m2.upper() == "EORM":
            continue
        codigo = str(ws["P2"].value or "").strip()
        texto = f"{titulo} — {m2}" + (f" ({codigo})" if codigo else "")
        escuelas.append((texto, titulo))
    wb.close()
    return escuelas


# =============================================================================
#  GENERADOR DE FICHAS SISCA
# =============================================================================

SISCA_HOJA_ADELANTE = "NOMINAL adelante"
SISCA_HOJA_ATRAS    = "NOMINAL atras"
SISCA_FILAS_POR_HOJA = 15
SISCA_ALUMNOS_POR_BLOQUE = 30
SISCA_FILA_INICIO_ADELANTE = 19
SISCA_FILA_INICIO_ATRAS    = 7

COLUMNAS_SUBTOTAL_SISCA = ("M", "N", "T")
FILA_SUBTOTAL_ATRAS = 6
FILA_TOTAL_ATRAS    = 22

def leer_alumnos_aptos_de_hoja(ws, fecha_corte: date, edad_min=6, edad_max=14) -> list:
    aptos = []
    for fila in range(FILA_INICIO_DATOS_ESCUELA, FILA_TOTAL_ESCUELA):
        nombre = ws.cell(fila, 3).value
        if not str(nombre or "").strip():
            continue
        if str(nombre).strip().startswith("---"):
            continue

        dia_nac  = ws.cell(fila, 16).value
        mes_nac  = ws.cell(fila, 17).value
        anio_nac = ws.cell(fila, 18).value
        edad = calcular_edad_a_fecha_corte(dia_nac, mes_nac, anio_nac, fecha_corte)
        if edad is None or not (edad_min <= edad <= edad_max):
            continue

        aptos.append({
            "nombre": str(nombre).strip(),
            "cui":    str(ws.cell(fila, 13).value or "").strip(),
            "genero": "F" if str(ws.cell(fila, 14).value or "").strip().upper() == "X" else
                      ("M" if str(ws.cell(fila, 15).value or "").strip().upper() == "X" else ""),
            "dia":  dia_nac,
            "mes":  mes_nac,
            "anio": anio_nac,
        })
    return aptos

def _clonar_imagenes(ws_origen, ws_destino, prefijo=""):
    imagenes_origen = getattr(ws_origen, "_images", None)
    if not imagenes_origen:
        return

    for img_original in imagenes_origen:
        try:
            pil_img = _import_image(img_original.ref)
            buf = BytesIO()
            fmt = getattr(pil_img, "format", None) or "PNG"
            pil_img.save(buf, format=fmt)
            buf.seek(0)

            nueva_imagen = XLImage(buf)
            nueva_imagen.width = img_original.width
            nueva_imagen.height = img_original.height
            nueva_imagen.anchor = copy.deepcopy(img_original.anchor)

            nombre_archivo = img_original.path.split("/")[-1]
            if prefijo:
                nueva_imagen._path = f"/xl/media/{prefijo}_{nombre_archivo}"
            else:
                nueva_imagen._path = f"/xl/media/clon_{id(nueva_imagen)}_{nombre_archivo}"
        except Exception:
            continue
        ws_destino.add_image(nueva_imagen)

def _fijar_formulas_subtotal_y_total(ws_atras, nombre_adelante: str):
    for col in COLUMNAS_SUBTOTAL_SISCA:
        celda_subtotal = f"{col}{FILA_SUBTOTAL_ATRAS}"
        formula_actual = ws_atras[celda_subtotal].value
        if isinstance(formula_actual, str) and formula_actual.startswith("="):
            if SISCA_HOJA_ADELANTE in formula_actual:
                ws_atras[celda_subtotal] = formula_actual.replace(f"'{SISCA_HOJA_ADELANTE}'", f"'{nombre_adelante}'")
        else:
            ws_atras[celda_subtotal] = f"='{nombre_adelante}'!{celda_subtotal}"

        celda_total = f"{col}{FILA_TOTAL_ATRAS}"
        rango_x = f"{col}{SISCA_FILA_INICIO_ATRAS}:{col}{FILA_TOTAL_ATRAS - 1}"
        ws_atras[celda_total] = f'=COUNTIF({rango_x},"X")+{celda_subtotal}'

def _asegurar_sheet_properties(ws):
    if ws is not None and ws.sheet_properties is None:
        from openpyxl.worksheet.properties import WorksheetProperties
        ws.sheet_properties = WorksheetProperties()

def _duplicar_bloque_sisca(wb, indice_bloque: int):
    if indice_bloque == 1:
        return wb[SISCA_HOJA_ADELANTE], wb[SISCA_HOJA_ATRAS]

    nombre_adelante = f"{SISCA_HOJA_ADELANTE} {indice_bloque}"
    nombre_atras    = f"{SISCA_HOJA_ATRAS} {indice_bloque}"

    ws_adelante_base = wb[SISCA_HOJA_ADELANTE]
    ws_atras_base    = wb[SISCA_HOJA_ATRAS]

    _asegurar_sheet_properties(ws_adelante_base)
    _asegurar_sheet_properties(ws_atras_base)

    ws_adelante = wb.copy_worksheet(ws_adelante_base)
    ws_adelante.title = nombre_adelante
    _clonar_imagenes(ws_adelante_base, ws_adelante, f"bloq{indice_bloque}")

    ws_atras = wb.copy_worksheet(ws_atras_base)
    ws_atras.title = nombre_atras
    _clonar_imagenes(ws_atras_base, ws_atras, f"bloq{indice_bloque}")

    try:
        if ws_adelante_base and ws_adelante:
            ws_adelante.print_area = ws_adelante_base.print_area
            if hasattr(ws_adelante_base, 'page_setup') and ws_adelante_base.page_setup:
                ws_adelante.page_setup.orientation = ws_adelante_base.page_setup.orientation
                ws_adelante.page_setup.paperSize = ws_adelante_base.page_setup.paperSize
                ws_adelante.page_setup.fitToPage = ws_adelante_base.page_setup.fitToPage
    except Exception as e:
        print(f"Nota: No se pudieron copiar dinámicamente todas las propiedades de impresión de adelante: {e}")

    try:
        if ws_atras_base and ws_atras:
            ws_atras.print_area = ws_atras_base.print_area
            if hasattr(ws_atras_base, 'page_setup') and ws_atras_base.page_setup:
                ws_atras.page_setup.orientation = ws_atras_base.page_setup.orientation
                ws_atras.page_setup.paperSize = ws_atras_base.page_setup.paperSize
                ws_atras.page_setup.fitToPage = ws_atras_base.page_setup.fitToPage
    except Exception as e:
        print(f"Nota: No se pudieron copiar dinámicamente todas las propiedades de impresión de atrás: {e}")

    _fijar_formulas_subtotal_y_total(ws_atras, nombre_adelante)
    return ws_adelante, ws_atras

def _aplicar_fuente_encabezado(celda):
    alineacion_previa = celda.alignment
    celda.font = Font(name="Arial", size=14, bold=True)
    celda.alignment = Alignment(
        horizontal=alineacion_previa.horizontal,
        vertical=alineacion_previa.vertical,
        wrap_text=alineacion_previa.wrap_text,
        text_rotation=alineacion_previa.text_rotation,
    )

def _rellenar_encabezado_sisca(ws_adelante, nombre_escuela: str, codigo_escuela: str,
                                anio_campana: int, tipo_centro: str = None,
                                ws_atras=None, area=None, distrito=None,
                                servicio=None, responsable=None, cargo=None,
                                fecha_reporte_str=None):
    tipo = (tipo_centro or TIPO_CENTRO_EDUCATIVO_DEFECTO).strip().upper()
    valores = {
        "C7":  area or "",
        "T7":  distrito or "",
        "AB7": servicio or "",
        "G9":  responsable or "",
        "S9":  cargo or "",
        "AI9": fecha_reporte_str or "",
        "AE9": "X",
        "E12": nombre_escuela.strip().upper(),
        "V12": codigo_escuela.strip().upper(),
    }
    if tipo == "PRIVADO":
        valores["AK12"] = "X"
    else:
        valores["AH12"] = "X"

    for direccion, valor in valores.items():
        ws_adelante[direccion] = valor
        _aplicar_fuente_encabezado(ws_adelante[direccion])

def _rellenar_alumnos_pagina(ws, fila_inicio: int, alumnos_pagina: list):
    for i in range(SISCA_FILAS_POR_HOJA):
        fila = fila_inicio + i
        for col in (2, 12, 13, 14, 15, 16, 17):
            ws.cell(fila, col, None)

    for i, alumno in enumerate(alumnos_pagina):
        fila = fila_inicio + i
        ws.cell(fila, 2, alumno["nombre"])
        celda_cui = ws.cell(fila, 12, alumno["cui"])
        celda_cui.number_format = "@"
        if alumno["genero"] == "F":
            ws.cell(fila, 13, "X")
        elif alumno["genero"] == "M":
            ws.cell(fila, 14, "X")
        ws.cell(fila, 15, alumno["dia"])
        ws.cell(fila, 16, alumno["mes"])
        ws.cell(fila, 17, alumno["anio"])

def generar_ficha_sisca_escuela(nombre_escuela: str, codigo_escuela: str,
                                 alumnos_aptos: list, anio_campana: int,
                                 tipo_centro: str = None,
                                 responsable=None, cargo=None,
                                 area=None, distrito=None, servicio=None,
                                 fecha_reporte_str=None) -> str:
    ruta_plantilla = ruta_plantilla_sisca()
    if not os.path.exists(ruta_plantilla):
        raise FileNotFoundError(
            f"No se encontró la plantilla legal en:\n{ruta_plantilla}\n\n"
            "Coloca el archivo 'plantilla_sisca.xlsx' junto al programa."
        )

    wb = openpyxl.load_workbook(ruta_plantilla)
    total_bloques = max(1, math.ceil(len(alumnos_aptos) / SISCA_ALUMNOS_POR_BLOQUE))
    bloques_hojas = [_duplicar_bloque_sisca(wb, indice) for indice in range(1, total_bloques + 1)]

    for indice, (ws_adelante, ws_atras) in enumerate(bloques_hojas, start=1):
        _rellenar_encabezado_sisca(ws_adelante, nombre_escuela, codigo_escuela,
                                    anio_campana, tipo_centro, ws_atras,
                                    area=area, distrito=distrito,
                                    servicio=servicio,
                                    responsable=responsable, cargo=cargo,
                                    fecha_reporte_str=fecha_reporte_str)

        inicio = (indice - 1) * SISCA_ALUMNOS_POR_BLOQUE
        bloque = alumnos_aptos[inicio:inicio + SISCA_ALUMNOS_POR_BLOQUE]
        pagina_adelante = bloque[:SISCA_FILAS_POR_HOJA]
        pagina_atras    = bloque[SISCA_FILAS_POR_HOJA:SISCA_ALUMNOS_POR_BLOQUE]

        _rellenar_alumnos_pagina(ws_adelante, SISCA_FILA_INICIO_ADELANTE, pagina_adelante)
        _rellenar_alumnos_pagina(ws_atras, SISCA_FILA_INICIO_ATRAS, pagina_atras)

    try:
        wb.calculation = CalcProperties(fullCalcOnLoad=True)
    except Exception:
        pass

    nombre_archivo = f"SISCA_{_nombre_archivo_seguro(codigo_escuela or nombre_escuela)}.xlsx"
    ruta_salida = os.path.join(carpeta_salida_sisca(), nombre_archivo)
    wb.save(ruta_salida)
    wb.close()
    gc.collect()
    return ruta_salida

def generar_sisca_para_hoja(nombre_hoja_escuela: str, fecha_corte: date,
                             responsable=None, cargo=None,
                             area=None, distrito=None, servicio=None,
                             tipo_centro=None, fecha_reporte_str=None) -> dict:
    wb = abrir_base_datos()
    if nombre_hoja_escuela not in wb.sheetnames:
        raise ValueError(f"La hoja '{nombre_hoja_escuela}' no existe en la Base de Datos.")

    ws = wb[nombre_hoja_escuela]
    nombre_escuela = str(ws["M2"].value or "").strip()
    codigo_escuela = str(ws["P2"].value or "").strip()
    wb.close()

    aptos = leer_alumnos_aptos_de_hoja(ws, fecha_corte)
    if not aptos:
        return {
            "hoja": nombre_hoja_escuela, "nombre": nombre_escuela,
            "codigo": codigo_escuela, "aptos": 0, "ruta": None,
        }

    ruta = generar_ficha_sisca_escuela(
        nombre_escuela, codigo_escuela, aptos, fecha_corte.year,
        tipo_centro=tipo_centro,
        responsable=responsable, cargo=cargo,
        area=area, distrito=distrito, servicio=servicio,
        fecha_reporte_str=fecha_reporte_str,
    )
    return {
        "hoja": nombre_hoja_escuela, "nombre": nombre_escuela,
        "codigo": codigo_escuela, "aptos": len(aptos), "ruta": ruta,
    }


# =============================================================================
#  UTILIDAD FLUORIZACIÓN INDEPENDIENTE
# =============================================================================

_NOMBRE_COL_INICIO = 2
_NOMBRE_COL_SPAN   = 10
_NOMBRE_COL_FIN    = _NOMBRE_COL_INICIO + _NOMBRE_COL_SPAN - 1

_COL_CUI_FLUO  = _NOMBRE_COL_FIN + 1
_COL_SEXO_FLUO = _COL_CUI_FLUO + 1
_COL_C2_FLUO   = _COL_SEXO_FLUO + 1
_COL_C11_FLUO  = _COL_C2_FLUO + 1
_COL_DIA_FLUO  = _COL_C11_FLUO + 1
_COL_MES_FLUO  = _COL_DIA_FLUO + 1
_COL_ANIO_FLUO = _COL_MES_FLUO + 1

_COLS_FLUORIZACION = (["No.", "Nombre"] + [""] * (_NOMBRE_COL_SPAN - 1) +
                      ["Cui", "Sexo", "Constante 2", "Constante 11",
                       "Día de nacimiento", "Mes de nacimiento", "Año de nacimiento"])
_ANCHOS_FLUORIZACION = ([6] + [4] * _NOMBRE_COL_SPAN + [16, 7, 12, 12, 18, 18, 18])
_COLOR_HEADER_FLUORIZACION = "1B2D9A"

def _borde_fino_fluo():
    lado = Side(style="thin")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _estilo_encabezado_fluo(ws, fila, columnas, color_fondo):
    font  = Font(bold=True, color="FFFFFF", size=11, name="Century Gothic")
    fill  = PatternFill("solid", start_color=color_fondo)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    borde = _borde_fino_fluo()
    for col_idx, nombre in enumerate(columnas, 1):
        c = ws.cell(row=fila, column=col_idx, value=nombre)
        c.font, c.fill, c.alignment, c.border = font, fill, align, borde
    ws.row_dimensions[fila].height = 28

def _estilo_separador_fluo(ws, fila, texto, num_cols, color_fondo):
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=num_cols)
    c = ws.cell(row=fila, column=1, value=texto)
    c.font = Font(bold=True, size=14, name="Arial", color="FFFFFF")
    c.fill = PatternFill("solid", start_color=color_fondo)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila].height = 22

def _estilo_dato_fluo(ws, fila, col_idx, valor):
    c = ws.cell(row=fila, column=col_idx, value=valor)
    c.font = Font(bold=True, size=14, name="Arial")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = _borde_fino_fluo()
    return c

def generar_fluorizacion(alumnos: list, ruta_salida: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fluorización"

    _estilo_encabezado_fluo(ws, 1, _COLS_FLUORIZACION, _COLOR_HEADER_FLUORIZACION)
    for i, ancho in enumerate(_ANCHOS_FLUORIZACION, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = ancho
    ws.merge_cells(start_row=1, start_column=_NOMBRE_COL_INICIO, end_row=1, end_column=_NOMBRE_COL_FIN)

    grado_ant = seccion_ant = None
    fila, corr = 2, 1

    for alumno in alumnos:
        g, s = alumno["grado"], alumno["seccion"]
        if g != grado_ant or s != seccion_ant:
            _estilo_separador_fluo(ws, fila, f"--- {g} - SECCIÓN {s} ---", len(_COLS_FLUORIZACION), _COLOR_HEADER_FLUORIZACION)
            fila += 1
            corr = 1
            grado_ant, seccion_ant = g, s

        dia, mes, anio = _split_fecha(alumno["fecha_nac"])
        nombre = _nombre_completo(alumno)
        sexo = "F" if alumno["genero"] == "FEMENINO" else "M"

        _estilo_dato_fluo(ws, fila, 1, corr)
        c_nombre = _estilo_dato_fluo(ws, fila, _NOMBRE_COL_INICIO, nombre)
        c_nombre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(_NOMBRE_COL_INICIO + 1, _NOMBRE_COL_FIN + 1):
            _estilo_dato_fluo(ws, fila, col, "")
        ws.merge_cells(start_row=fila, start_column=_NOMBRE_COL_INICIO, end_row=fila, end_column=_NOMBRE_COL_FIN)

        c_cui = _estilo_dato_fluo(ws, fila, _COL_CUI_FLUO, alumno["cui"])
        c_cui.number_format = "@"
        _estilo_dato_fluo(ws, fila, _COL_SEXO_FLUO, sexo)
        _estilo_dato_fluo(ws, fila, _COL_C2_FLUO, 2)
        _estilo_dato_fluo(ws, fila, _COL_C11_FLUO, 11)
        _estilo_dato_fluo(ws, fila, _COL_DIA_FLUO, dia)
        _estilo_dato_fluo(ws, fila, _COL_MES_FLUO, mes)
        c_anio = _estilo_dato_fluo(ws, fila, _COL_ANIO_FLUO, anio)
        c_anio.alignment = Alignment(horizontal="center", vertical="center", text_rotation=90)

        fila += 1
        corr += 1
    wb.save(ruta_salida)

def procesar_pdfs_individual(tipo: str, parent=None):
    rutas_pdf, _ = QFileDialog.getOpenFileNames(
        parent, "Seleccionar uno o varios PDFs", "",
        "Archivos PDF (*.pdf)"
    )
    if not rutas_pdf:
        return

    generados, errores = [], []
    for ruta_pdf in rutas_pdf:
        carpeta = os.path.dirname(ruta_pdf)
        nombre_pdf = os.path.splitext(os.path.basename(ruta_pdf))[0]
        ruta_excel = os.path.join(carpeta, f"{nombre_pdf}_{tipo}.xlsx")
        try:
            alumnos = extraer_alumnos_pdf(ruta_pdf)
            if not alumnos:
                errores.append(f"{os.path.basename(ruta_pdf)}: sin alumnos encontrados")
                continue
            generar_fluorizacion(alumnos, ruta_excel)
            generados.append(f"{os.path.basename(ruta_excel)}  ({len(alumnos)} alumnos)")
        except Exception as e:
            errores.append(f"{os.path.basename(ruta_pdf)}: {e}")

    msg = ""
    if generados:
        msg += "✅ Archivos generados:\n" + "\n".join(f"  • {g}" for g in generados)
    if errores:
        msg += "\n\n❌ Errores:\n" + "\n".join(f"  • {e}" for e in errores)
    if errores and not generados:
        QMessageBox.critical(parent, "Error", msg)
    elif errores:
        QMessageBox.warning(parent, "Completado con advertencias", msg)
    else:
        QMessageBox.information(parent, "Listo ✓", msg)


# =============================================================================
#  INTERFAZ GRÁFICA — PyQt6 GLASSMORPHISM REAL
# =============================================================================

ANCHO_VENTANA = 1000
ALTO_VENTANA  = 700

COLOR_FONDO_VENTANA  = "#050810"
COLOR_PANEL_FONDO    = "#0D1B2A"
COLOR_TITULO         = "#FFFFFF"
COLOR_SUBTITULO      = "#A0B8D8"
COLOR_ETIQUETA       = "#B0C8E8"
COLOR_PIE            = "#6080A0"
COLOR_FIRMA          = "#FFFFFF"

COLOR_BTN_INSERTAR   = "#2E6FE0"
COLOR_BTN_INSERTAR_2 = "#123B8A"
COLOR_BTN_GENERAR    = "#2FC7C2"
COLOR_BTN_GENERAR_2  = "#12707A"
COLOR_BTN_FLUOR      = "#2FA8D8"
COLOR_BTN_FLUOR_2    = "#125E82"

COLOR_TAB_SEL        = "#2A56B8"
FUENTE_TITULO_SERIF = "Georgia"
FUENTE_INTERFAZ     = "Century Gothic"

ANCHO_MINIMO = 820
ALTO_MINIMO  = 620

PANEL_ANCHO   = 560
PANEL_ALTO    = 430
PANEL_RADIUS  = 25
BTN_RADIUS    = 18


def _cargar_fondo_qpixmap(ruta, ancho, alto):
    """Carga y escala la imagen de fondo para que cubra toda la ventana (cover)."""
    pix = QPixmap(ruta)
    if pix.isNull():
        return QPixmap(ancho, alto)
    scaled = pix.scaled(ancho, alto,
                        Qt.AspectRatioMode.KeepAspectRatioByExpanding,
                        Qt.TransformationMode.SmoothTransformation)
    x = (scaled.width() - ancho) // 2
    y = (scaled.height() - alto) // 2
    return scaled.copy(x, y, ancho, alto)


class DialogoSisca(QDialog):
    def __init__(self, config_previa: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Datos del Encabezado SISCA")
        self.setFixedSize(520, 420)
        self.setModal(True)

        self.setStyleSheet("""
            QDialog {
                background-color: rgba(10, 21, 37, 0.95);
                border: 2px solid rgba(255, 255, 255, 0.6);
                border-radius: 20px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 18, 24, 18)
        layout.setSpacing(10)

        titulo = QLabel("Encabezado de la Ficha SISCA")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet(
            "font: bold 16px 'Georgia'; color: white; background: transparent;")
        layout.addWidget(titulo)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        estilo_label = (
            "font: bold 11px 'Century Gothic'; color: #B0C8E8; background: transparent;")
        estilo_edit = (
            "font: 12px 'Century Gothic'; color: #B0C8E8;"
            "background-color: rgba(10, 21, 37, 0.8);"
            "border: 1px solid #2A4A78; border-radius: 8px; padding: 5px 8px;")

        self.responsable_entry = QLineEdit(config_previa.get("responsable", ""))
        self.responsable_entry.setPlaceholderText("Nombre del responsable")
        self.responsable_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Responsable:"), self.responsable_entry)
        form.labelForField(self.responsable_entry).setStyleSheet(estilo_label)

        self.cargo_entry = QLineEdit(config_previa.get("cargo", ""))
        self.cargo_entry.setPlaceholderText("Ej: Técnico en Salud Rural")
        self.cargo_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Cargo:"), self.cargo_entry)
        form.labelForField(self.cargo_entry).setStyleSheet(estilo_label)

        self.area_entry = QLineEdit(config_previa.get("area_salud", ""))
        self.area_entry.setPlaceholderText("Ej: TOTONICAPÁN")
        self.area_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Área de Salud:"), self.area_entry)
        form.labelForField(self.area_entry).setStyleSheet(estilo_label)

        self.distrito_entry = QLineEdit(config_previa.get("distrito_salud", ""))
        self.distrito_entry.setPlaceholderText("Ej: TOTONICAPÁN")
        self.distrito_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Distrito de Salud:"), self.distrito_entry)
        form.labelForField(self.distrito_entry).setStyleSheet(estilo_label)

        self.servicio_entry = QLineEdit(config_previa.get("servicio_salud", ""))
        self.servicio_entry.setPlaceholderText("Ej: Aldea Chuanoj")
        self.servicio_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Servicio de Salud:"), self.servicio_entry)
        form.labelForField(self.servicio_entry).setStyleSheet(estilo_label)

        hoy = date.today()
        fecha_defecto = f"{hoy.day:02d}/{hoy.month:02d}/{hoy.year}"
        self.fecha_entry = QLineEdit(fecha_defecto)
        self.fecha_entry.setStyleSheet(estilo_edit)
        form.addRow(QLabel("Fecha del Reporte:"), self.fecha_entry)
        form.labelForField(self.fecha_entry).setStyleSheet(estilo_label)

        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems(["PÚBLICO", "PRIVADO"])
        idx = self.tipo_combo.findText(
            config_previa.get("tipo_centro", "PÚBLICO").upper())
        if idx >= 0:
            self.tipo_combo.setCurrentIndex(idx)
        self.tipo_combo.setStyleSheet("""
            QComboBox {
                background-color: rgba(10, 21, 37, 0.8);
                border: 1px solid #2A4A78; border-radius: 8px;
                color: #B0C8E8; padding: 5px 8px;
                font: 12px 'Century Gothic';
            }
            QComboBox::drop-down {
                border: none; border-radius: 8px;
                background: #2A4A78; width: 22px;
            }
            QComboBox::down-arrow { image: none; }
            QComboBox QAbstractItemView {
                background: #0A1525; color: #B0C8E8;
                selection-background-color: #2A56B8;
                border: 1px solid #2A4A78; border-radius: 6px;
            }
        """)
        form.addRow(QLabel("Tipo Centro:"), self.tipo_combo)
        form.labelForField(self.tipo_combo).setStyleSheet(estilo_label)

        layout.addLayout(form)
        layout.addStretch()

        btn_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                                   QDialogButtonBox.StandardButton.Cancel)
        btn_box.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2E6FE0, stop:1 #123B8A);
                border-radius: 14px; color: white;
                font: bold 12px 'Century Gothic'; padding: 8px 22px;
                min-width: 100px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3A80F0, stop:1 #1A4A9A);
            }
        """)
        btn_box.accepted.connect(self._confirmar)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

    def _confirmar(self):
        if not self.responsable_entry.text().strip():
            QMessageBox.warning(self, "Campo obligatorio",
                "Escribe el nombre del Responsable de la Información.")
            return
        if not self.fecha_entry.text().strip():
            QMessageBox.warning(self, "Campo obligatorio",
                "Escribe la fecha del reporte.")
            return
        datos = {
            "responsable": self.responsable_entry.text().strip(),
            "cargo": self.cargo_entry.text().strip(),
            "area_salud": self.area_entry.text().strip(),
            "distrito_salud": self.distrito_entry.text().strip(),
            "servicio_salud": self.servicio_entry.text().strip(),
            "tipo_centro": self.tipo_combo.currentText().strip(),
            "fecha_reporte": self.fecha_entry.text().strip(),
        }
        guardar_config_tsr(datos)
        self._datos = datos
        self.accept()

    @property
    def datos(self) -> dict:
        return getattr(self, "_datos", {})


class InterfazTSR(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("TSR - Herramienta de Salud Rural")
        self.resize(ANCHO_VENTANA, ALTO_VENTANA)
        self.setMinimumSize(ANCHO_MINIMO, ALTO_MINIMO)

        screen = QApplication.primaryScreen().geometry()
        self.move((screen.width() - ANCHO_VENTANA) // 2,
                  (screen.height() - ALTO_VENTANA) // 2)

        central = QWidget()
        self.setCentralWidget(central)

        # ── Fondo con mapa nítido ──────────────────────────────────────
        self.bg_label = QLabel(central)
        ruta_fondo = _obtener_ruta_recurso(NOMBRES_ARCHIVOS["fondo"])
        if os.path.exists(ruta_fondo):
            self._bg_pixmap = _cargar_fondo_qpixmap(ruta_fondo, ANCHO_VENTANA, ALTO_VENTANA)
            self.bg_label.setPixmap(self._bg_pixmap)
        self.bg_label.setGeometry(0, 0, ANCHO_VENTANA, ALTO_VENTANA)

        # ── Menú central de vidrio (glassmorphism real con QSS) ────────
        self.glass_frame = QFrame(central)
        self.glass_frame.setObjectName("glassFrame")
        self.glass_frame.setStyleSheet("""
            QFrame#glassFrame {
                background-color: rgba(13, 27, 42, 0.45);
                border: 2px solid rgba(255, 255, 255, 0.7);
                border-radius: 25px;
            }
        """)
        self.glass_frame.setGeometry(0, 0, PANEL_ANCHO, PANEL_ALTO)
        px = (ANCHO_VENTANA - PANEL_ANCHO) // 2
        py = (ALTO_VENTANA - PANEL_ALTO) // 2
        self.glass_frame.move(px, py)

        glass_layout = QVBoxLayout(self.glass_frame)
        glass_layout.setContentsMargins(20, 15, 20, 15)
        glass_layout.setSpacing(0)

        # ── Título ─────────────────────────────────────────────────────
        lbl_titulo = QLabel("TSR")
        lbl_titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_titulo.setStyleSheet(
            f"font: bold 36px '{FUENTE_TITULO_SERIF}'; color: {COLOR_TITULO}; background: transparent;")
        glass_layout.addWidget(lbl_titulo)

        lbl_sub = QLabel("HERRAMIENTA DE SALUD RURAL")
        lbl_sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_sub.setStyleSheet(
            f"font: bold 10px '{FUENTE_INTERFAZ}'; color: {COLOR_SUBTITULO}; background: transparent;")
        glass_layout.addWidget(lbl_sub)

        # ── Pestañas ───────────────────────────────────────────────────
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { background: transparent; border: none; }
            QTabBar::tab {
                background: rgba(10, 21, 37, 0.7);
                color: white;
                padding: 6px 18px;
                border-top-left-radius: 12px;
                border-top-right-radius: 12px;
                font: bold 11px 'Century Gothic';
                margin-right: 2px;
            }
            QTabBar::tab:selected { background: #2A56B8; }
            QTabBar::tab:hover:!selected { background: #1A3A7A; }
        """)
        glass_layout.addWidget(self.tab_widget, 1)

        tab_despar = QWidget()
        tab_despar.setStyleSheet("background: transparent;")
        tab_fluor = QWidget()
        tab_fluor.setStyleSheet("background: transparent;")

        self.tab_widget.addTab(tab_despar, "   \u2695  Desparasitación   ")
        self.tab_widget.addTab(tab_fluor, "   \U0001F9B7  Fluorización   ")

        self._mapa_escuelas = {}

        self._construir_tab_desparasitacion(tab_despar)
        self._construir_tab_fluorizacion(tab_fluor)

        # ── Pie ───────────────────────────────────────────────────────
        pie_layout = QHBoxLayout()
        pie_layout.setContentsMargins(0, 0, 0, 0)
        lbl_pie = QLabel(
            "La Base de Datos y las Fichas SISCA se generan junto al programa")
        lbl_pie.setStyleSheet(
            f"font: 9px '{FUENTE_INTERFAZ}'; color: {COLOR_PIE}; background: transparent;")
        pie_layout.addWidget(lbl_pie)
        pie_layout.addStretch()
        lbl_firma = QLabel("TSR Romeo Mazariegos")
        lbl_firma.setStyleSheet(
            f"font: 8px '{FUENTE_INTERFAZ}'; color: {COLOR_FIRMA}; background: transparent;")
        pie_layout.addWidget(lbl_firma)
        glass_layout.addLayout(pie_layout)

        # ── Escuelas ───────────────────────────────────────────────────
        self._refrescar_combobox_escuelas()

        # ── Icono ──────────────────────────────────────────────────────
        ruta_ico = _obtener_ruta_recurso(NOMBRES_ARCHIVOS["icono"])
        if os.path.exists(ruta_ico):
            try:
                from PyQt6.QtGui import QIcon
                self.setWindowIcon(QIcon(ruta_ico))
            except Exception:
                pass

    # -- Operaciones de negocio --------------------------------------------

    def _obtener_fecha_corte(self):
        try:
            d = int(self.dia_entry.text().strip())
            m = int(self.mes_entry.text().strip())
            a = int(self.anio_entry.text().strip())
            return date(a, m, d)
        except Exception:
            QMessageBox.critical(self, "Fecha de corte inválida",
                "Escribe una fecha de corte válida en los campos Día / Mes / Año.")
            return None

    def _pedir_dato_manual(self, mensaje: str):
        text, ok = QInputDialog.getText(self, "Dato no detectado", mensaje)
        return text if ok else ""

    def _refrescar_combobox_escuelas(self, seleccionar_hoja=None):
        try:
            escuelas = listar_escuelas_procesadas()
        except Exception:
            escuelas = []
        self._mapa_escuelas = {texto: hoja for texto, hoja in escuelas}
        textos = list(self._mapa_escuelas.keys())
        self.combo_escuelas.clear()
        self.combo_escuelas.addItems(textos)
        if seleccionar_hoja:
            for texto, hoja in self._mapa_escuelas.items():
                if hoja == seleccionar_hoja:
                    idx = self.combo_escuelas.findText(texto)
                    if idx >= 0:
                        self.combo_escuelas.setCurrentIndex(idx)
                    return
        if textos and not self.combo_escuelas.currentText():
            self.combo_escuelas.setCurrentIndex(0)

    def accion_insertar_escuela(self):
        fecha_corte = self._obtener_fecha_corte()
        if fecha_corte is None:
            return
        rutas_pdf, _ = QFileDialog.getOpenFileNames(
            self, "Seleccionar uno o varios listados PDF de escuelas", "",
            "Archivos PDF (*.pdf)")
        if not rutas_pdf:
            return
        resultados = []
        errores = []
        for ruta_pdf in rutas_pdf:
            try:
                resultado = insertar_nueva_escuela_desde_pdf(
                    ruta_pdf, fecha_corte, pedir_dato_manual=self._pedir_dato_manual)
                resultados.append(resultado)
            except Exception as e:
                errores.append(f"{os.path.basename(ruta_pdf)}: {e}")
        gc.collect()
        if resultados:
            self._refrescar_combobox_escuelas(seleccionar_hoja=resultados[-1]["hoja"])
        msg = ""
        if resultados:
            msg += "Procesados:\n" + "\n".join(
                f"  \u2022 {r['hoja']} \u2014 {r['nombre_completo'][:50]} ({r['agregados']} alumnos)"
                for r in resultados)
        if errores:
            msg += "\n\nErrores:\n" + "\n".join(f"  \u2022 {e}" for e in errores)
        if not resultados:
            QMessageBox.critical(self, "Error",
                msg if msg else "No se procesó ningún archivo.")
        elif errores:
            QMessageBox.warning(self, "Completado con advertencias", msg)
        else:
            QMessageBox.information(self, "Listo \u2713", msg)

    def accion_generar_sisca_escuela(self):
        fecha_corte = self._obtener_fecha_corte()
        if fecha_corte is None:
            return
        texto_sel = self.combo_escuelas.currentText().strip()
        hoja = self._mapa_escuelas.get(texto_sel)
        if not hoja:
            QMessageBox.warning(self, "Selecciona una escuela",
                "Primero inserta o selecciona una escuela procesada en el combobox.")
            return

        config = cargar_config_tsr()
        dialogo = DialogoSisca(config, self)
        if dialogo.exec() != QDialog.DialogCode.Accepted:
            return
        d = dialogo.datos

        try:
            resultado = generar_sisca_para_hoja(
                hoja, fecha_corte,
                responsable=d.get("responsable", ""),
                cargo=d.get("cargo", ""),
                area=d.get("area_salud", ""),
                distrito=d.get("distrito_salud", ""),
                servicio=d.get("servicio_salud", ""),
                tipo_centro=d.get("tipo_centro", "PÚBLICO"),
                fecha_reporte_str=d.get("fecha_reporte", ""),
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
            return
        if not resultado["ruta"]:
            QMessageBox.warning(self, "Sin alumnos aptos",
                f"{resultado['hoja']} ({resultado['nombre']}) no tiene alumnos de 6 a 14 años.")
            return
        QMessageBox.information(self, "Listo ✓",
            f"Ficha SISCA generada para {resultado['nombre']} ({resultado['aptos']} alumnos):\n\n{resultado['ruta']}")

    def accion_fluorizacion(self):
        procesar_pdfs_individual("fluorizacion", parent=self)

    # -- Construcción de tabs ---------------------------------------------

    def _construir_tab_desparasitacion(self, tab):
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 5, 0, 0)
        layout.setSpacing(8)

        fila_fecha = QHBoxLayout()
        fila_fecha.setSpacing(6)

        lbl_fecha = QLabel("Fecha de corte:")
        lbl_fecha.setStyleSheet(
            f"font: bold 12px '{FUENTE_INTERFAZ}'; color: {COLOR_ETIQUETA}; background: transparent;")
        fila_fecha.addWidget(lbl_fecha)
        fila_fecha.addSpacing(4)

        entry_style = (
            f"font: bold 14px '{FUENTE_INTERFAZ}'; color: #B0C8E8;"
            "background-color: rgba(10, 21, 37, 0.8);"
            "border: 1px solid #2A4A78; border-radius: 10px; padding: 4px 6px;")

        self.dia_entry = QLineEdit(f"{FECHA_CORTE_DEFECTO[0]:02d}")
        self.dia_entry.setFixedWidth(36)
        self.dia_entry.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.dia_entry.setStyleSheet(entry_style)
        fila_fecha.addWidget(self.dia_entry)

        sep = QLabel("/")
        sep.setStyleSheet(
            "font: bold 14px 'Century Gothic'; color: #6080A0; background: transparent;")
        fila_fecha.addWidget(sep)

        self.mes_entry = QLineEdit(f"{FECHA_CORTE_DEFECTO[1]:02d}")
        self.mes_entry.setFixedWidth(36)
        self.mes_entry.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.mes_entry.setStyleSheet(entry_style)
        fila_fecha.addWidget(self.mes_entry)

        sep2 = QLabel("/")
        sep2.setStyleSheet(
            "font: bold 14px 'Century Gothic'; color: #6080A0; background: transparent;")
        fila_fecha.addWidget(sep2)

        self.anio_entry = QLineEdit(str(FECHA_CORTE_DEFECTO[2]))
        self.anio_entry.setFixedWidth(60)
        self.anio_entry.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.anio_entry.setStyleSheet(entry_style)
        fila_fecha.addWidget(self.anio_entry)

        fila_fecha.addStretch()
        layout.addLayout(fila_fecha)
        layout.addSpacing(4)

        btn_ins = QPushButton("+  Insertar Escuela (PDF)")
        btn_ins.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_ins.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {COLOR_BTN_INSERTAR}, stop:1 {COLOR_BTN_INSERTAR_2});
                border-radius: {BTN_RADIUS}px; color: white;
                font: bold 13px '{FUENTE_INTERFAZ}'; padding: 12px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3A80F0, stop:1 #1A4A9A);
            }}
        """)
        btn_ins.clicked.connect(self.accion_insertar_escuela)
        layout.addWidget(btn_ins)
        layout.addSpacing(6)

        lbl_esc = QLabel("Escuela procesada:")
        lbl_esc.setStyleSheet(
            f"font: bold 12px '{FUENTE_INTERFAZ}'; color: {COLOR_ETIQUETA}; background: transparent;")
        layout.addWidget(lbl_esc)

        self.combo_escuelas = QComboBox()
        self.combo_escuelas.setStyleSheet("""
            QComboBox {
                background-color: rgba(10, 21, 37, 0.8);
                border: 1px solid #2A4A78; border-radius: 10px;
                color: #B0C8E8; padding: 6px 10px;
                font: 11px 'Century Gothic';
            }
            QComboBox::drop-down {
                border: none; border-radius: 10px;
                background: #2A4A78; width: 24px;
            }
            QComboBox::down-arrow { image: none; }
            QComboBox QAbstractItemView {
                background: #0A1525; color: #B0C8E8;
                selection-background-color: #2A56B8;
                border: 1px solid #2A4A78; border-radius: 6px;
            }
        """)
        layout.addWidget(self.combo_escuelas)
        layout.addSpacing(6)

        btn_gen = QPushButton("\U0001F4CB  Generar Ficha SISCA")
        btn_gen.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_gen.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {COLOR_BTN_GENERAR}, stop:1 {COLOR_BTN_GENERAR_2});
                border-radius: {BTN_RADIUS}px; color: white;
                font: bold 13px '{FUENTE_INTERFAZ}'; padding: 12px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3AD8D3, stop:1 #1A808A);
            }}
        """)
        btn_gen.clicked.connect(self.accion_generar_sisca_escuela)
        layout.addWidget(btn_gen)
        layout.addStretch()

    def _construir_tab_fluorizacion(self, tab):
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(0, 5, 0, 0)
        layout.setSpacing(8)

        lbl_desc = QLabel(
            "Genera un listado de Fluorización independiente\n(no toca la Base de Datos Central).")
        lbl_desc.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_desc.setStyleSheet(
            f"font: bold 12px '{FUENTE_INTERFAZ}'; color: {COLOR_ETIQUETA}; background: transparent;")
        layout.addWidget(lbl_desc)
        layout.addSpacing(16)

        btn_fluor = QPushButton("\U0001F9B7  Generar Listado Fluorización")
        btn_fluor.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_fluor.setStyleSheet(f"""
            QPushButton {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {COLOR_BTN_FLUOR}, stop:1 {COLOR_BTN_FLUOR_2});
                border-radius: {BTN_RADIUS}px; color: white;
                font: bold 13px '{FUENTE_INTERFAZ}'; padding: 12px;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3FB8E8, stop:1 #1A6E92);
            }}
        """)
        btn_fluor.clicked.connect(self.accion_fluorizacion)
        layout.addWidget(btn_fluor)
        layout.addStretch()

    # -- Redimensión -------------------------------------------------------

    def resizeEvent(self, event):
        super().resizeEvent(event)
        w = event.size().width()
        h = event.size().height()
        ruta_fondo = _obtener_ruta_recurso(NOMBRES_ARCHIVOS["fondo"])
        if os.path.exists(ruta_fondo):
            self._bg_pixmap = _cargar_fondo_qpixmap(ruta_fondo, w, h)
            self.bg_label.setPixmap(self._bg_pixmap)
        self.bg_label.setGeometry(0, 0, w, h)
        px = (w - PANEL_ANCHO) // 2
        py = (h - PANEL_ALTO) // 2
        self.glass_frame.move(px, py)


# =============================================================================
#  PUNTO DE ENTRADA
# =============================================================================

if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ventana = InterfazTSR()
    ventana.show()
    sys.exit(app.exec())